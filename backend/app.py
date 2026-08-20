import base64
import csv
import io
import logging
import os
import re
import sqlite3
import uuid
from datetime import datetime
from logging.handlers import RotatingFileHandler
from pathlib import Path

import mammoth
from flask import Flask, abort, jsonify, request, send_file, send_from_directory
from flask_cors import CORS
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent

DB_PATH = Path(
    os.environ.get("DATABASE_PATH", BASE_DIR / "customer_remark.db")
).expanduser()
LOG_PATH = Path(os.environ.get("LOG_PATH", BASE_DIR / "malstar_toolkit.log")).expanduser()
UPLOAD_DIR = Path(os.environ.get("UPLOAD_DIR", BASE_DIR / "uploads")).expanduser()
MAX_UPLOAD_MB = int(os.environ.get("MAX_UPLOAD_MB", "32"))
MAX_JSON_UPLOAD_MB = int(os.environ.get("MAX_JSON_UPLOAD_MB", "4"))
PREVIEW_ROWS = 200
PREVIEW_COLS = 30
ALLOWED_FILE_KINDS = {
    ".docx": "docx",
    ".xlsx": "xlsx",
}
CORS_ORIGINS = [
    origin.strip()
    for origin in os.environ.get(
        "CORS_ORIGINS", "http://localhost:5173,http://127.0.0.1:5173"
    ).split(",")
    if origin.strip()
]


def resolve_static_dir():
    env_path = os.environ.get("STATIC_DIR")
    if env_path:
        return Path(env_path)
    candidates = [
        BASE_DIR / "frontend" / "dist",
        BASE_DIR.parent / "frontend" / "dist",
    ]
    for path in candidates:
        if path.exists():
            return path
    return candidates[0]


STATIC_DIR = resolve_static_dir()

app = Flask(__name__, static_folder=None)
app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_MB * 1024 * 1024
app.config["MAX_FORM_MEMORY_SIZE"] = MAX_UPLOAD_MB * 1024 * 1024
app.config["MAX_FORM_PARTS"] = 10000
if CORS_ORIGINS:
    CORS(app, resources={r"/api/*": {"origins": CORS_ORIGINS}})


def now_stamp():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def setup_file_logger():
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("malstar")
    if logger.handlers:
        return logger
    logger.setLevel(logging.INFO)
    handler = RotatingFileHandler(LOG_PATH, maxBytes=5 * 1024 * 1024, backupCount=5, encoding="utf-8")
    handler.setFormatter(logging.Formatter("%(message)s"))
    logger.addHandler(handler)
    logger.propagate = False
    return logger


APP_LOGGER = setup_file_logger()


def log_event(action, detail=""):
    stamp = now_stamp()
    try:
        client_ip = request.remote_addr or ""
    except RuntimeError:
        client_ip = ""
    detail = str(detail or "").strip()
    line = f"{stamp} | {action}"
    if detail:
        line += f" | {detail}"
    if client_ip:
        line += f" | ip={client_ip}"
    APP_LOGGER.info(line)
    try:
        with get_connection() as conn:
            conn.execute(
                """
                INSERT INTO ActivityLogs (Timestamp, Action, Detail, ClientIP)
                VALUES (?, ?, ?, ?)
                """,
                (stamp, action[:200], detail[:2000], client_ip[:80]),
            )
    except sqlite3.Error:
        pass
    return stamp


def letters_only(value):
    return "".join(ch for ch in (value or "") if ch.isalpha()).casefold()


def record_values(payload):
    return (
        payload["ctrlOrgcode"],
        payload["customer"],
        payload["customerLetters"],
        payload["remark1"],
        payload["remark2"],
        payload["remark3"],
    )


def get_connection():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=5000")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db():
    with get_connection() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS CustomerRemarks (
                ID INTEGER PRIMARY KEY AUTOINCREMENT,
                CTRLOrgcode TEXT NOT NULL,
                Customer TEXT NOT NULL,
                CustomerLetters TEXT NOT NULL DEFAULT '',
                Remark1 TEXT NOT NULL DEFAULT '',
                Remark2 TEXT NOT NULL DEFAULT '',
                Remark3 TEXT NOT NULL DEFAULT '',
                CreateTime TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UpdateTime TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE (CTRLOrgcode, Customer)
            )
        """)
        columns = {
            row[1] for row in conn.execute("PRAGMA table_info(CustomerRemarks)")
        }
        if "CustomerLetters" not in columns:
            conn.execute(
                "ALTER TABLE CustomerRemarks "
                "ADD COLUMN CustomerLetters TEXT NOT NULL DEFAULT ''"
            )
        for row in conn.execute("SELECT ID, Customer FROM CustomerRemarks"):
            conn.execute(
                "UPDATE CustomerRemarks SET CustomerLetters=? WHERE ID=?",
                (letters_only(row["Customer"]), row["ID"]),
            )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_customer_remarks_orgcode "
            "ON CustomerRemarks (CTRLOrgcode)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_customer_remarks_customer "
            "ON CustomerRemarks (Customer)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_customer_remarks_letters "
            "ON CustomerRemarks (CustomerLetters)"
        )
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ActivityLogs (
                ID INTEGER PRIMARY KEY AUTOINCREMENT,
                Timestamp TEXT NOT NULL,
                Action TEXT NOT NULL,
                Detail TEXT NOT NULL DEFAULT '',
                ClientIP TEXT NOT NULL DEFAULT ''
            )
        """)
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_activity_logs_time ON ActivityLogs (Timestamp DESC, ID DESC)"
        )
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ToolkitFiles (
                ID INTEGER PRIMARY KEY AUTOINCREMENT,
                OriginalName TEXT NOT NULL,
                StoredName TEXT NOT NULL UNIQUE,
                Kind TEXT NOT NULL,
                Size INTEGER NOT NULL DEFAULT 0,
                UploadedAt TEXT NOT NULL,
                UpdatedAt TEXT NOT NULL
            )
        """)
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_toolkit_files_name ON ToolkitFiles (OriginalName)"
        )
        conn.execute("""
            CREATE TABLE IF NOT EXISTS Sops (
                ID INTEGER PRIMARY KEY AUTOINCREMENT,
                Title TEXT NOT NULL,
                Purpose TEXT NOT NULL DEFAULT '',
                Owner TEXT NOT NULL DEFAULT '',
                Revision TEXT NOT NULL DEFAULT '',
                Status TEXT NOT NULL DEFAULT 'draft',
                CreatedAt TEXT NOT NULL,
                UpdatedAt TEXT NOT NULL
            )
        """)
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_sops_title ON Sops (Title)"
        )
        conn.execute("""
            CREATE TABLE IF NOT EXISTS SopSteps (
                ID INTEGER PRIMARY KEY AUTOINCREMENT,
                SopID INTEGER NOT NULL,
                StepNumber INTEGER NOT NULL,
                Instruction TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (SopID) REFERENCES Sops(ID) ON DELETE CASCADE
            )
        """)
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_sop_steps_sop ON SopSteps (SopID, StepNumber)"
        )
        conn.execute("""
            CREATE TABLE IF NOT EXISTS SopAttachments (
                SopID INTEGER NOT NULL,
                FileID INTEGER NOT NULL,
                PRIMARY KEY (SopID, FileID),
                FOREIGN KEY (SopID) REFERENCES Sops(ID) ON DELETE CASCADE,
                FOREIGN KEY (FileID) REFERENCES ToolkitFiles(ID)
            )
        """)
        count = conn.execute("SELECT COUNT(*) FROM CustomerRemarks").fetchone()[0]
        if count == 0:
            seeds = [
                ("CQN", "Demo Customer A", "Priority customer", "Weekly review", "Active"),
                ("SHA", "Demo Customer B", "Standard process", "", "Active"),
                ("HKG", "Demo Customer C", "Check instruction", "Confirm before release", ""),
            ]
            conn.executemany(
                """
                INSERT INTO CustomerRemarks
                    (CTRLOrgcode, Customer, CustomerLetters, Remark1, Remark2, Remark3)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                [
                    (org, customer, letters_only(customer), remark1, remark2, remark3)
                    for org, customer, remark1, remark2, remark3 in seeds
                ],
            )


init_db()
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
log_event("MALSTAR_Toolkit started", f"database={DB_PATH}")


def parse_payload(data):
    ctrl_orgcode = str(data.get("ctrlOrgcode", data.get("CTRLOrgcode", "")) or "").strip().upper()
    customer = str(data.get("customer", data.get("Customer", "")) or "").strip()
    if not ctrl_orgcode:
        return None, "CTRLOrgcode is required."
    if not customer:
        return None, "Customer is required."
    return {
        "ctrlOrgcode": ctrl_orgcode,
        "customer": customer,
        "customerLetters": letters_only(customer),
        "remark1": str(data.get("remark1", data.get("Remark1", "")) or "").strip(),
        "remark2": str(data.get("remark2", data.get("Remark2", "")) or "").strip(),
        "remark3": str(data.get("remark3", data.get("Remark3", "")) or "").strip(),
    }, None


def row_to_dict(row):
    return {
        "id": row["ID"],
        "ctrlOrgcode": row["CTRLOrgcode"],
        "customer": row["Customer"],
        "remark1": row["Remark1"],
        "remark2": row["Remark2"],
        "remark3": row["Remark3"],
        "createTime": row["CreateTime"],
        "updateTime": row["UpdateTime"],
    }


def file_to_dict(row):
    return {
        "id": row["ID"],
        "originalName": row["OriginalName"],
        "kind": row["Kind"],
        "size": row["Size"],
        "uploadedAt": row["UploadedAt"],
        "updatedAt": row["UpdatedAt"],
    }


def ensure_upload_dir():
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    return UPLOAD_DIR


def file_kind_from_name(filename):
    suffix = Path(filename or "").suffix.lower()
    return ALLOWED_FILE_KINDS.get(suffix), suffix


def stored_path(stored_name):
    path = (ensure_upload_dir() / stored_name).resolve()
    if path.parent != ensure_upload_dir().resolve():
        abort(404)
    return path


def save_bytes_as_file(original_name, data):
    kind, suffix = file_kind_from_name(original_name)
    if not kind:
        return None, "Please upload a .docx or .xlsx file."
    if not data:
        return None, "The uploaded file is empty."
    if len(data) > MAX_UPLOAD_MB * 1024 * 1024:
        return None, f"File exceeds the {MAX_UPLOAD_MB} MB limit."
    display_name = Path(original_name).name.strip() or f"upload{suffix}"
    stored_name = f"{uuid.uuid4().hex}{suffix}"
    dest = stored_path(stored_name)
    dest.write_bytes(data)
    stamp = now_stamp()
    with get_connection() as conn:
        cur = conn.execute(
            """
            INSERT INTO ToolkitFiles (OriginalName, StoredName, Kind, Size, UploadedAt, UpdatedAt)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (display_name, stored_name, kind, len(data), stamp, stamp),
        )
        row = conn.execute("SELECT * FROM ToolkitFiles WHERE ID=?", (cur.lastrowid,)).fetchone()
    return file_to_dict(row), None


def cell_to_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def sanitize_preview_html(value):
    cleaned = re.sub(r"(?is)<script.*?>.*?</script>", "", value or "")
    cleaned = re.sub(r"(?is)<iframe.*?>.*?</iframe>", "", cleaned)
    return cleaned


def preview_docx(path):
    with path.open("rb") as handle:
        result = mammoth.convert_to_html(handle)
    return {
        "kind": "docx",
        "html": sanitize_preview_html(result.value),
        "messages": [str(item) for item in (result.messages or [])][:8],
    }


def preview_xlsx(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    try:
        for worksheet in workbook.worksheets:
            rows = []
            truncated = False
            for row_index, row in enumerate(worksheet.iter_rows(max_col=PREVIEW_COLS, values_only=True), start=1):
                if row_index > PREVIEW_ROWS:
                    truncated = True
                    break
                values = [cell_to_text(cell) for cell in row]
                if any(values):
                    rows.append(values)
            sheets.append({
                "name": worksheet.title,
                "rows": rows,
                "truncated": truncated,
            })
    finally:
        workbook.close()
    return {"kind": "xlsx", "sheets": sheets}


def load_sop(conn, sop_id):
    row = conn.execute("SELECT * FROM Sops WHERE ID=?", (sop_id,)).fetchone()
    if not row:
        return None
    steps = conn.execute(
        """
        SELECT ID, StepNumber, Instruction
        FROM SopSteps WHERE SopID=? ORDER BY StepNumber, ID
        """,
        (sop_id,),
    ).fetchall()
    attachments = conn.execute(
        """
        SELECT f.* FROM ToolkitFiles f
        INNER JOIN SopAttachments a ON a.FileID = f.ID
        WHERE a.SopID=?
        ORDER BY f.OriginalName
        """,
        (sop_id,),
    ).fetchall()
    return {
        "id": row["ID"],
        "title": row["Title"],
        "purpose": row["Purpose"],
        "owner": row["Owner"],
        "revision": row["Revision"],
        "status": row["Status"],
        "createdAt": row["CreatedAt"],
        "updatedAt": row["UpdatedAt"],
        "steps": [
            {
                "id": step["ID"],
                "stepNumber": step["StepNumber"],
                "instruction": step["Instruction"],
            }
            for step in steps
        ],
        "attachments": [file_to_dict(item) for item in attachments],
    }


def parse_sop_payload(data):
    title = str(data.get("title") or "").strip()
    if not title:
        return None, "Title is required."
    status = str(data.get("status") or "draft").strip().lower()
    if status not in {"draft", "active"}:
        status = "draft"
    raw_steps = data.get("steps") if isinstance(data.get("steps"), list) else []
    steps = []
    for item in raw_steps:
        if isinstance(item, str):
            instruction = item.strip()
        elif isinstance(item, dict):
            instruction = str(item.get("instruction") or "").strip()
        else:
            instruction = ""
        if instruction:
            steps.append(instruction)
    raw_ids = data.get("attachmentIds") if isinstance(data.get("attachmentIds"), list) else []
    attachment_ids = []
    for item in raw_ids:
        try:
            file_id = int(item)
        except (TypeError, ValueError):
            continue
        if file_id not in attachment_ids:
            attachment_ids.append(file_id)
    return {
        "title": title[:200],
        "purpose": str(data.get("purpose") or "").strip(),
        "owner": str(data.get("owner") or "").strip()[:120],
        "revision": str(data.get("revision") or "").strip()[:40],
        "status": status,
        "steps": steps,
        "attachmentIds": attachment_ids,
    }, None


def replace_sop_children(conn, sop_id, payload):
    conn.execute("DELETE FROM SopSteps WHERE SopID=?", (sop_id,))
    conn.execute("DELETE FROM SopAttachments WHERE SopID=?", (sop_id,))
    for index, instruction in enumerate(payload["steps"], start=1):
        conn.execute(
            "INSERT INTO SopSteps (SopID, StepNumber, Instruction) VALUES (?, ?, ?)",
            (sop_id, index, instruction),
        )
    for file_id in payload["attachmentIds"]:
        exists = conn.execute("SELECT ID FROM ToolkitFiles WHERE ID=?", (file_id,)).fetchone()
        if not exists:
            continue
        conn.execute(
            "INSERT INTO SopAttachments (SopID, FileID) VALUES (?, ?)",
            (sop_id, file_id),
        )


def upsert_imported_records(records):
    created = updated = 0
    with get_connection() as conn:
        for payload in records:
            existing = conn.execute(
                """
                SELECT ID FROM CustomerRemarks
                WHERE CTRLOrgcode=? AND Customer=?
                """,
                (payload["ctrlOrgcode"], payload["customer"]),
            ).fetchone()
            if existing:
                updated += 1
                conn.execute(
                    """
                    UPDATE CustomerRemarks
                    SET CustomerLetters=?, Remark1=?, Remark2=?, Remark3=?,
                        UpdateTime=CURRENT_TIMESTAMP WHERE ID=?
                    """,
                    (
                        payload["customerLetters"],
                        payload["remark1"],
                        payload["remark2"],
                        payload["remark3"],
                        existing["ID"],
                    ),
                )
            else:
                created += 1
                conn.execute(
                    """
                    INSERT INTO CustomerRemarks
                        (CTRLOrgcode, Customer, CustomerLetters, Remark1, Remark2, Remark3)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    record_values(payload),
                )
    return created, updated


def collect_import_payloads(raw_rows):
    records_by_key, errors, duplicates = {}, [], 0
    for row_no, raw in enumerate(raw_rows, start=2):
        if not any(str(v or "").strip() for v in raw.values()):
            continue
        payload, error = parse_payload(raw)
        if error:
            errors.append({"row": row_no, "message": error})
            continue
        key = (payload["ctrlOrgcode"].casefold(), payload["customer"].casefold())
        if key in records_by_key:
            duplicates += 1
        records_by_key[key] = payload
    return list(records_by_key.values()), errors, duplicates


@app.get("/api/health")
def health():
    return jsonify({"success": True, "message": "MALSTAR_Toolkit API is running"})


@app.get("/api/activity-logs")
def list_activity_logs():
    page = max(request.args.get("page", 1, type=int), 1)
    page_size = min(max(request.args.get("pageSize", 50, type=int), 1), 200)
    with get_connection() as conn:
        total = conn.execute("SELECT COUNT(*) FROM ActivityLogs").fetchone()[0]
        rows = conn.execute(
            """
            SELECT ID, Timestamp, Action, Detail, ClientIP
            FROM ActivityLogs
            ORDER BY ID DESC
            LIMIT ? OFFSET ?
            """,
            (page_size, (page - 1) * page_size),
        ).fetchall()
    return jsonify({
        "success": True,
        "data": [
            {
                "id": row["ID"],
                "timestamp": row["Timestamp"],
                "action": row["Action"],
                "detail": row["Detail"],
                "clientIp": row["ClientIP"],
            }
            for row in rows
        ],
        "pagination": {
            "page": page,
            "pageSize": page_size,
            "total": total,
            "totalPages": max((total + page_size - 1) // page_size, 1),
        },
    })


@app.post("/api/activity-logs")
def create_activity_log():
    data = request.get_json(silent=True) or {}
    action = str(data.get("action") or "").strip() or "UI event"
    detail = str(data.get("detail") or "").strip()
    stamp = log_event(action, detail)
    return jsonify({"success": True, "timestamp": stamp})


@app.get("/api/files")
def list_files():
    q = str(request.args.get("q") or "").strip()
    with get_connection() as conn:
        if q:
            rows = conn.execute(
                """
                SELECT * FROM ToolkitFiles
                WHERE OriginalName LIKE ?
                ORDER BY UploadedAt DESC, ID DESC
                """,
                (f"%{q}%",),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM ToolkitFiles ORDER BY UploadedAt DESC, ID DESC"
            ).fetchall()
    return jsonify({"success": True, "data": [file_to_dict(row) for row in rows]})


def store_uploaded_file(filename, data):
    record, error = save_bytes_as_file(filename, data)
    if error:
        log_event("File upload failed", error)
        return jsonify({"success": False, "message": error}), 400
    log_event("File uploaded", f"{record['originalName']} ({record['kind']})")
    return jsonify({"success": True, "message": "File uploaded", "data": record}), 201


@app.post("/api/files")
def upload_file():
    if "file" in request.files:
        file = request.files["file"]
        filename = file.filename or ""
        data = file.read()
        return store_uploaded_file(filename, data)
    data = request.get_json(silent=True) or {}
    filename = str(data.get("filename") or "")
    raw = str(data.get("content") or data.get("contentBase64") or "")
    if not raw:
        log_event("File upload failed", "no file uploaded")
        return jsonify({"success": False, "message": "No file was uploaded."}), 400
    try:
        payload = base64.b64decode(raw, validate=False)
    except Exception:
        return jsonify({"success": False, "message": "The file data is not valid base64."}), 400
    if len(payload) > MAX_JSON_UPLOAD_MB * 1024 * 1024:
        return jsonify({
            "success": False,
            "message": f"JSON uploads are limited to {MAX_JSON_UPLOAD_MB} MB. Use a smaller file.",
        }), 400
    return store_uploaded_file(filename, payload)


@app.get("/api/files/<int:file_id>/preview")
def preview_file(file_id):
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM ToolkitFiles WHERE ID=?", (file_id,)).fetchone()
    if not row:
        return jsonify({"success": False, "message": "File not found"}), 404
    path = stored_path(row["StoredName"])
    if not path.is_file():
        return jsonify({"success": False, "message": "File is missing on disk"}), 404
    try:
        preview = preview_docx(path) if row["Kind"] == "docx" else preview_xlsx(path)
    except Exception as error:
        log_event("File preview failed", f"id={file_id} {error}")
        return jsonify({"success": False, "message": f"Could not preview this file. {error}"}), 400
    preview["file"] = file_to_dict(row)
    return jsonify({"success": True, "data": preview})


@app.get("/api/files/<int:file_id>")
def download_file(file_id):
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM ToolkitFiles WHERE ID=?", (file_id,)).fetchone()
    if not row:
        return jsonify({"success": False, "message": "File not found"}), 404
    path = stored_path(row["StoredName"])
    if not path.is_file():
        return jsonify({"success": False, "message": "File is missing on disk"}), 404
    return send_file(path, as_attachment=True, download_name=row["OriginalName"])


@app.patch("/api/files/<int:file_id>")
def rename_file(file_id):
    data = request.get_json(silent=True) or {}
    name = Path(str(data.get("originalName") or data.get("name") or "")).name.strip()
    if not name:
        return jsonify({"success": False, "message": "A file name is required."}), 400
    kind, suffix = file_kind_from_name(name)
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM ToolkitFiles WHERE ID=?", (file_id,)).fetchone()
        if not row:
            return jsonify({"success": False, "message": "File not found"}), 404
        if kind and kind != row["Kind"]:
            name = f"{Path(name).stem}{Path(row['StoredName']).suffix}"
        elif not Path(name).suffix:
            name = f"{name}{Path(row['StoredName']).suffix}"
        conn.execute(
            "UPDATE ToolkitFiles SET OriginalName=?, UpdatedAt=? WHERE ID=?",
            (name, now_stamp(), file_id),
        )
        updated = conn.execute("SELECT * FROM ToolkitFiles WHERE ID=?", (file_id,)).fetchone()
    log_event("File renamed", f"id={file_id} {name}")
    return jsonify({"success": True, "message": "File renamed", "data": file_to_dict(updated)})


@app.delete("/api/files/<int:file_id>")
def delete_file(file_id):
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM ToolkitFiles WHERE ID=?", (file_id,)).fetchone()
        if not row:
            return jsonify({"success": False, "message": "File not found"}), 404
        attached = conn.execute(
            "SELECT SopID FROM SopAttachments WHERE FileID=? LIMIT 1",
            (file_id,),
        ).fetchone()
        if attached:
            return jsonify({
                "success": False,
                "message": "This file is attached to an SOP. Detach it first, then delete.",
            }), 409
        conn.execute("DELETE FROM ToolkitFiles WHERE ID=?", (file_id,))
    path = stored_path(row["StoredName"])
    if path.is_file():
        path.unlink()
    log_event("File deleted", f"id={file_id} {row['OriginalName']}")
    return jsonify({"success": True, "message": "File deleted"})


@app.get("/api/sops")
def list_sops():
    q = str(request.args.get("q") or "").strip()
    with get_connection() as conn:
        if q:
            rows = conn.execute(
                """
                SELECT * FROM Sops
                WHERE Title LIKE ? OR Owner LIKE ? OR Revision LIKE ?
                ORDER BY UpdatedAt DESC, ID DESC
                """,
                (f"%{q}%", f"%{q}%", f"%{q}%"),
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM Sops ORDER BY UpdatedAt DESC, ID DESC").fetchall()
        data = []
        for row in rows:
            step_count = conn.execute(
                "SELECT COUNT(*) FROM SopSteps WHERE SopID=?", (row["ID"],)
            ).fetchone()[0]
            file_count = conn.execute(
                "SELECT COUNT(*) FROM SopAttachments WHERE SopID=?", (row["ID"],)
            ).fetchone()[0]
            data.append({
                "id": row["ID"],
                "title": row["Title"],
                "purpose": row["Purpose"],
                "owner": row["Owner"],
                "revision": row["Revision"],
                "status": row["Status"],
                "createdAt": row["CreatedAt"],
                "updatedAt": row["UpdatedAt"],
                "stepCount": step_count,
                "attachmentCount": file_count,
            })
    return jsonify({"success": True, "data": data})


@app.post("/api/sops")
def create_sop():
    payload, error = parse_sop_payload(request.get_json(silent=True) or {})
    if error:
        log_event("SOP create failed", error)
        return jsonify({"success": False, "message": error}), 400
    stamp = now_stamp()
    with get_connection() as conn:
        cur = conn.execute(
            """
            INSERT INTO Sops (Title, Purpose, Owner, Revision, Status, CreatedAt, UpdatedAt)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                payload["title"],
                payload["purpose"],
                payload["owner"],
                payload["revision"],
                payload["status"],
                stamp,
                stamp,
            ),
        )
        sop_id = cur.lastrowid
        replace_sop_children(conn, sop_id, payload)
        data = load_sop(conn, sop_id)
    log_event("SOP created", payload["title"])
    return jsonify({"success": True, "message": "SOP created", "data": data}), 201


@app.get("/api/sops/<int:sop_id>")
def get_sop(sop_id):
    with get_connection() as conn:
        data = load_sop(conn, sop_id)
    if not data:
        return jsonify({"success": False, "message": "SOP not found"}), 404
    return jsonify({"success": True, "data": data})


@app.put("/api/sops/<int:sop_id>")
def update_sop(sop_id):
    payload, error = parse_sop_payload(request.get_json(silent=True) or {})
    if error:
        log_event("SOP update failed", error)
        return jsonify({"success": False, "message": error}), 400
    with get_connection() as conn:
        existing = conn.execute("SELECT ID FROM Sops WHERE ID=?", (sop_id,)).fetchone()
        if not existing:
            return jsonify({"success": False, "message": "SOP not found"}), 404
        conn.execute(
            """
            UPDATE Sops
            SET Title=?, Purpose=?, Owner=?, Revision=?, Status=?, UpdatedAt=?
            WHERE ID=?
            """,
            (
                payload["title"],
                payload["purpose"],
                payload["owner"],
                payload["revision"],
                payload["status"],
                now_stamp(),
                sop_id,
            ),
        )
        replace_sop_children(conn, sop_id, payload)
        data = load_sop(conn, sop_id)
    log_event("SOP updated", f"id={sop_id} {payload['title']}")
    return jsonify({"success": True, "message": "SOP saved", "data": data})


@app.delete("/api/sops/<int:sop_id>")
def delete_sop(sop_id):
    with get_connection() as conn:
        row = conn.execute("SELECT Title FROM Sops WHERE ID=?", (sop_id,)).fetchone()
        if not row:
            return jsonify({"success": False, "message": "SOP not found"}), 404
        conn.execute("DELETE FROM SopSteps WHERE SopID=?", (sop_id,))
        conn.execute("DELETE FROM SopAttachments WHERE SopID=?", (sop_id,))
        conn.execute("DELETE FROM Sops WHERE ID=?", (sop_id,))
    log_event("SOP deleted", f"id={sop_id} {row['Title']}")
    return jsonify({"success": True, "message": "SOP deleted"})


@app.get("/api/customer-remarks")
def list_records():
    q_letters = letters_only(request.args.get("q", ""))
    page = max(request.args.get("page", 1, type=int), 1)
    page_size = min(max(request.args.get("pageSize", 20, type=int), 1), 10000)
    if q_letters:
        where = """WHERE CustomerLetters != '' AND (
                       instr(CustomerLetters, ?) > 0
                       OR instr(?, CustomerLetters) > 0
                   )"""
        params = [q_letters, q_letters]
    else:
        where = ""
        params = []
    with get_connection() as conn:
        total = conn.execute(
            f"SELECT COUNT(*) FROM CustomerRemarks {where}", params
        ).fetchone()[0]
        rows = conn.execute(
            f"""
            SELECT * FROM CustomerRemarks {where}
            ORDER BY UpdateTime DESC, ID DESC LIMIT ? OFFSET ?
            """,
            params + [page_size, (page - 1) * page_size],
        ).fetchall()
    if q_letters:
        log_event("Search", f"query={request.args.get('q', '')} matches={total} page={page}")
    elif page > 1:
        log_event("List records", f"page={page} total={total}")
    return jsonify({
        "success": True,
        "data": [row_to_dict(row) for row in rows],
        "pagination": {
            "page": page,
            "pageSize": page_size,
            "total": total,
            "totalPages": max((total + page_size - 1) // page_size, 1),
        },
    })


@app.post("/api/customer-remarks")
def create_record():
    payload, error = parse_payload(request.get_json(silent=True) or {})
    if error:
        log_event("Create failed", error)
        return jsonify({"success": False, "message": error}), 400
    try:
        with get_connection() as conn:
            cur = conn.execute(
                """
                INSERT INTO CustomerRemarks
                    (CTRLOrgcode, Customer, CustomerLetters, Remark1, Remark2, Remark3)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                record_values(payload),
            )
            row = conn.execute(
                "SELECT * FROM CustomerRemarks WHERE ID=?", (cur.lastrowid,)
            ).fetchone()
    except sqlite3.IntegrityError:
        log_event(
            "Create failed",
            f"duplicate {payload['ctrlOrgcode']} / {payload['customer']}",
        )
        return jsonify({
            "success": False,
            "message": "The CTRLOrgcode and Customer combination already exists.",
        }), 409
    log_event("Record created", f"{payload['ctrlOrgcode']} / {payload['customer']}")
    return jsonify({"success": True, "message": "Record created", "data": row_to_dict(row)}), 201


@app.put("/api/customer-remarks/<int:record_id>")
def update_record(record_id):
    payload, error = parse_payload(request.get_json(silent=True) or {})
    if error:
        log_event("Update failed", error)
        return jsonify({"success": False, "message": error}), 400
    try:
        with get_connection() as conn:
            cur = conn.execute(
                """
                UPDATE CustomerRemarks
                SET CTRLOrgcode=?, Customer=?, CustomerLetters=?,
                    Remark1=?, Remark2=?, Remark3=?,
                    UpdateTime=CURRENT_TIMESTAMP
                WHERE ID=?
                """,
                record_values(payload) + (record_id,),
            )
            if cur.rowcount == 0:
                log_event("Update failed", f"id={record_id} not found")
                return jsonify({"success": False, "message": "Record not found"}), 404
            row = conn.execute(
                "SELECT * FROM CustomerRemarks WHERE ID=?", (record_id,)
            ).fetchone()
    except sqlite3.IntegrityError:
        log_event(
            "Update failed",
            f"duplicate {payload['ctrlOrgcode']} / {payload['customer']}",
        )
        return jsonify({
            "success": False,
            "message": "The CTRLOrgcode and Customer combination already exists.",
        }), 409
    log_event("Record updated", f"id={record_id} {payload['ctrlOrgcode']} / {payload['customer']}")
    return jsonify({"success": True, "message": "Record updated", "data": row_to_dict(row)})


@app.delete("/api/customer-remarks/<int:record_id>")
def delete_record(record_id):
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM CustomerRemarks WHERE ID=?", (record_id,)).fetchone()
        cur = conn.execute("DELETE FROM CustomerRemarks WHERE ID=?", (record_id,))
    if cur.rowcount == 0:
        log_event("Delete failed", f"id={record_id} not found")
        return jsonify({"success": False, "message": "Record not found"}), 404
    log_event(
        "Record deleted",
        f"id={record_id} {row['CTRLOrgcode']} / {row['Customer']}",
    )
    return jsonify({"success": True, "message": "Record deleted"})


@app.post("/api/customer-remarks/import")
def import_records():
    data = request.get_json(silent=True) or {}
    filename = str(data.get("filename") or "browser.csv")
    raw_rows = data.get("records")
    if not isinstance(raw_rows, list):
        log_event("CSV import failed", "JSON records missing")
        return jsonify({"success": False, "message": "No records were sent."}), 400
    log_event("CSV import started", f"file={filename} jsonRows={len(raw_rows)}")
    records, errors, duplicates = collect_import_payloads(raw_rows)
    if errors:
        log_event("CSV import failed", f"file={filename} validation errors={len(errors)}")
        return jsonify({
            "success": False,
            "message": "CSV validation failed. Nothing was imported.",
            "errors": errors[:100],
        }), 400
    if not records:
        log_event("CSV import failed", f"file={filename} no data rows")
        return jsonify({"success": False, "message": "CSV contains no data rows."}), 400
    created, updated = upsert_imported_records(records)
    log_event(
        "CSV import completed",
        f"file={filename} processed={len(records)} created={created} updated={updated} duplicates={duplicates}",
    )
    return jsonify({
        "success": True,
        "message": "CSV import completed",
        "processed": len(records),
        "created": created,
        "updated": updated,
        "duplicates": duplicates,
    })


@app.post("/api/customer-remarks/import-csv")
def import_csv():
    if "file" not in request.files:
        log_event("CSV import failed", "no file uploaded")
        return jsonify({"success": False, "message": "No CSV file was uploaded."}), 400
    file = request.files["file"]
    filename = file.filename or ""
    log_event("CSV import started", f"file={filename}")
    if not file.filename or not file.filename.lower().endswith(".csv"):
        log_event("CSV import failed", f"file={filename} not a csv")
        return jsonify({"success": False, "message": "Please select a .csv file."}), 400
    try:
        reader = csv.DictReader(io.TextIOWrapper(file.stream, encoding="utf-8-sig", newline=""))
        if not reader.fieldnames:
            return jsonify({"success": False, "message": "CSV header row is missing."}), 400
        headers = {str(h).strip().lower(): h for h in reader.fieldnames if h}
        required = ["ctrlorgcode", "customer"]
        missing = [h for h in required if h not in headers]
        if missing:
            return jsonify({
                "success": False,
                "message": "Missing CSV column(s): " + ", ".join(missing),
            }), 400

        def value(row, name):
            original = headers.get(name.lower())
            return row.get(original, "") if original else ""

        raw_rows = []
        for row in reader:
            if not any(str(v or "").strip() for v in row.values()):
                continue
            raw_rows.append({
                "ctrlOrgcode": value(row, "ctrlorgcode"),
                "customer": value(row, "customer"),
                "remark1": value(row, "remark1"),
                "remark2": value(row, "remark2"),
                "remark3": value(row, "remark3"),
            })
        records, errors, duplicates = collect_import_payloads(raw_rows)

        if errors:
            log_event("CSV import failed", f"file={filename} validation errors={len(errors)}")
            return jsonify({
                "success": False,
                "message": "CSV validation failed. Nothing was imported.",
                "errors": errors[:100],
            }), 400
        if not records:
            log_event("CSV import failed", f"file={filename} no data rows")
            return jsonify({"success": False, "message": "CSV contains no data rows."}), 400

        created, updated = upsert_imported_records(records)
        log_event(
            "CSV import completed",
            f"file={filename} processed={len(records)} created={created} updated={updated} duplicates={duplicates}",
        )
        return jsonify({
            "success": True,
            "message": "CSV import completed",
            "processed": len(records),
            "created": created,
            "updated": updated,
            "duplicates": duplicates,
        })
    except UnicodeDecodeError:
        log_event("CSV import failed", f"file={filename} encoding error")
        return jsonify({
            "success": False,
            "message": "Please save the CSV with UTF-8 encoding.",
        }), 400
    except csv.Error as error:
        log_event("CSV import failed", f"file={filename} {error}")
        return jsonify({"success": False, "message": f"Invalid CSV: {error}"}), 400


@app.errorhandler(413)
def too_large(_):
    return jsonify({
        "success": False,
        "message": f"Upload exceeds the {MAX_UPLOAD_MB} MB limit.",
    }), 413


@app.get("/")
def index():
    if not (STATIC_DIR / "index.html").is_file():
        return jsonify({
            "success": True,
            "message": "MALSTAR_Toolkit API is running. Frontend build not found.",
        })
    return send_from_directory(STATIC_DIR, "index.html")


@app.get("/<path:asset_path>")
def spa_or_static(asset_path):
    if asset_path.startswith("api/"):
        abort(404)
    target = STATIC_DIR / asset_path
    if target.is_file():
        return send_from_directory(STATIC_DIR, asset_path)
    if (STATIC_DIR / "index.html").is_file():
        return send_from_directory(STATIC_DIR, "index.html")
    abort(404)


@app.after_request
def write_request_log(response):
    path = request.path or ""
    if (
        path.startswith("/api/")
        and path not in {"/api/activity-logs", "/api/health"}
        and request.method in {"POST", "PUT", "PATCH", "DELETE"}
    ):
        log_event(
            f"{request.method} {path}",
            f"status={response.status_code}",
        )
    return response


if __name__ == "__main__":
    host = os.environ.get("FLASK_HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", os.environ.get("FLASK_PORT", "5000")))
    debug = os.environ.get("FLASK_DEBUG", "true").lower() in ("1", "true", "yes")
    app.run(host=host, port=port, debug=debug)
