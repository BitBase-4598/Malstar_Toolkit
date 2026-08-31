import os
import json
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
TMP = Path(tempfile.mkdtemp())
os.environ["DATABASE_PATH"] = str(TMP / "test.db")
os.environ["UPLOAD_DIR"] = str(TMP / "uploads")
os.environ["LOG_PATH"] = str(TMP / "test.log")
os.environ["GCA_XLSX_PATH"] = str(TMP / "missing-gca.xlsx")
os.environ["LCL_XLSX_PATH"] = str(TMP / "missing-lcl.xlsx")
os.environ["UNLOCODE_CSV_PATH"] = str(TMP / "missing-unlocode.csv")
sys.path.insert(0, str(ROOT))

from flask import Flask
from werkzeug.exceptions import NotFound

from db import get_connection, migrate
from services.dashboard_analytics import format_minutes, minutes_between, parse_dashboard_record
from services.files_store import stored_path
from util import letters_only


def test_letters_only():
    assert letters_only("Maersk-Line 123") == "maerskline"
    assert letters_only("") == ""


def test_format_minutes():
    assert format_minutes(None) == ""
    assert format_minutes(17) == "17 min"
    assert format_minutes(120) == "2h"
    assert format_minutes(235) == "3h 55m"


def test_parse_dashboard_record():
    record, error = parse_dashboard_record({
        "orderNumber": "OD1",
        "date": "2026-08-21",
        "emailReceived": "2026-08-21 08:00:00",
        "handlingTime": "2026-08-21 08:10:00",
        "bookingConvertedTime": "2026-08-21 08:40:00",
    })
    assert error is None
    assert record["reportDate"] == "2026-08-21"
    assert record["processMinutes"] == 30
    assert minutes_between(
        __import__("datetime").datetime(2026, 8, 21, 8, 0, 0),
        __import__("datetime").datetime(2026, 8, 21, 8, 10, 0),
    ) == 10


def test_migrate_schema_version_once():
    migrate()
    with get_connection() as conn:
        version = conn.execute("SELECT Version FROM SchemaVersion WHERE ID=1").fetchone()[0]
        count = conn.execute("SELECT COUNT(*) FROM CustomerRemarks").fetchone()[0]
    assert version == 5
    assert count == 3
    migrate()
    with get_connection() as conn:
        count_again = conn.execute("SELECT COUNT(*) FROM CustomerRemarks").fetchone()[0]
    assert count_again == 3


def test_stored_path_rejects_traversal():
    migrate()
    flask_app = Flask(__name__)
    with flask_app.app_context():
        try:
            stored_path("../secret.txt")
            raised = False
        except NotFound:
            raised = True
    assert raised


def test_api_health_leave_dashboard():
    from app import app

    client = app.test_client()
    health = client.get("/api/health")
    assert health.status_code == 200
    assert health.get_json()["success"] is True
    people = client.get("/api/leave-people")
    assert people.status_code == 200
    leave = client.get("/api/leave-plans?year=2026&month=8")
    assert leave.status_code == 200
    dashboard = client.get("/api/dashboard")
    assert dashboard.status_code == 200
    assert "kpis" in dashboard.get_json()["data"]
    lcl_filters = client.get("/api/lcl/filters")
    assert lcl_filters.status_code == 200
    assert "years" in lcl_filters.get_json()["data"]
    lcl_dash = client.get("/api/lcl/dashboard")
    assert lcl_dash.status_code == 200
    lcl_data = lcl_dash.get_json()["data"]
    assert "summary" in lcl_data
    assert "map" in lcl_data
    icb = client.get("/api/icb")
    assert icb.status_code == 200
    unloco = client.get("/api/unlocode")
    assert unloco.status_code == 200
    gca = client.get("/api/gca/summary")
    assert gca.status_code == 200
    assert "kpis" in gca.get_json()["data"]


def test_audit_writes_sqlite_and_json():
    migrate()
    from config import LOG_PATH
    from logging_util import audit

    audit("record.create", summary="CQN / Demo Customer A", resource_id="1")
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM ActivityLogs ORDER BY ID DESC LIMIT 1").fetchone()
    assert row["ActionCode"] == "record.create"
    assert row["Outcome"] == "success"
    assert row["Module"] == "records"
    assert "T" in row["Timestamp"] and row["Timestamp"].endswith("Z")
    lines = Path(LOG_PATH).read_text(encoding="utf-8").strip().splitlines()
    payload = json.loads(lines[-1])
    assert payload["event"] == "record.create"
    assert payload["outcome"] == "success"
    assert payload["level"] == "INFO"


def test_activity_log_failure_filter():
    migrate()
    from logging_util import audit
    from app import app

    audit("record.create", "failure", summary="duplicate key")
    audit("record.create", summary="ok row")
    client = app.test_client()
    data = client.get("/api/activity-logs?outcome=failure").get_json()["data"]
    assert data
    assert all(item["outcome"] in ("failure", "exception") for item in data)
    assert any(item["detail"] == "duplicate key" for item in data)


def test_activity_log_action_filter_returns_200():
    migrate()
    from logging_util import audit
    from app import app

    audit("record.create", summary="CQN / Demo Customer A", resource_id="1")
    client = app.test_client()
    response = client.get("/api/activity-logs?action=record")
    assert response.status_code == 200
    body = response.get_json()
    assert body["success"] is True
    assert any("record" in str(item.get("action") or "").lower() or "record" in str(item.get("actionCode") or "").lower() for item in body["data"])


def test_get_list_does_not_audit():
    migrate()
    from app import app

    with get_connection() as conn:
        before = conn.execute("SELECT COUNT(*) FROM ActivityLogs").fetchone()[0]
    client = app.test_client()
    response = client.get("/api/customer-remarks")
    assert response.status_code == 200
    with get_connection() as conn:
        after = conn.execute("SELECT COUNT(*) FROM ActivityLogs").fetchone()[0]
    assert after == before


def test_unhandled_api_exception_is_audited():
    migrate()
    from app import create_app

    test_app = create_app()

    def boom():
        raise RuntimeError("boom-test")

    test_app.add_url_rule("/api/__test-boom", endpoint="test_boom_exc", view_func=boom)
    client = test_app.test_client()
    response = client.get("/api/__test-boom")
    assert response.status_code == 500
    body = response.get_json()
    assert body["success"] is False
    assert response.headers.get("X-Request-ID")
    with get_connection() as conn:
        row = conn.execute(
            "SELECT * FROM ActivityLogs WHERE ActionCode='server.exception' ORDER BY ID DESC LIMIT 1"
        ).fetchone()
    assert row is not None
    assert row["Outcome"] == "exception"
    assert "boom-test" in row["Summary"]


def test_cases_crud_and_status():
    migrate()
    from app import app

    client = app.test_client()
    missing = client.post("/api/cases", json={"name": "Only name"})
    assert missing.status_code == 400
    created = client.post("/api/cases", json={
        "category": "Human Error",
        "description": "POR was read as Guangzhou",
    })
    assert created.status_code == 201
    body = created.get_json()["data"]
    assert body["status"] == "pending_review"
    assert body["category"] == "Human Error"
    assert body["description"] == "POR was read as Guangzhou"
    assert body["hbl"] == ""
    case_id = body["id"]
    listed = client.get("/api/cases?q=Guangzhou").get_json()["data"]
    assert any(item["id"] == case_id for item in listed)
    patched = client.patch(f"/api/cases/{case_id}/status", json={"status": "closed"})
    assert patched.status_code == 200
    assert patched.get_json()["data"]["status"] == "closed"
    deleted = client.delete(f"/api/cases/{case_id}")
    assert deleted.status_code == 200
    missing_get = client.get(f"/api/cases/{case_id}")
    assert missing_get.status_code == 404


def test_case_file_upload_png_xlsx():
    from io import BytesIO

    from openpyxl import Workbook

    migrate()
    from app import app

    client = app.test_client()
    created = client.post("/api/cases", json={
        "category": "Defects",
        "description": "Attach demo",
    }).get_json()["data"]
    case_id = created["id"]
    png = bytes.fromhex(
        "89504e470d0a1a0a0000000d49484452000000010000000108060000001f15c489"
        "0000000a49444154789c63000100000500010d0a2db40000000049454e44ae426082"
    )
    image = client.post(
        f"/api/cases/{case_id}/files",
        data={"file": (BytesIO(png), "screenshot.png")},
        content_type="multipart/form-data",
    )
    assert image.status_code == 201
    assert image.get_json()["data"]["kind"] == "image"
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "HBL"
    sheet["B1"] = "HBL-FILE"
    buffer = BytesIO()
    workbook.save(buffer)
    excel = client.post(
        f"/api/cases/{case_id}/files",
        data={"file": (BytesIO(buffer.getvalue()), "correction.xlsx")},
        content_type="multipart/form-data",
    )
    assert excel.status_code == 201
    assert excel.get_json()["data"]["kind"] == "xlsx"
    detail = client.get(f"/api/cases/{case_id}").get_json()["data"]
    assert detail["fileCount"] == 2
    with get_connection() as conn:
        library = conn.execute("SELECT COUNT(*) FROM ToolkitFiles").fetchone()[0]
    assert library == 0


def test_case_import_and_locked_update():
    from io import BytesIO

    migrate()
    from app import app

    client = app.test_client()
    csv_bytes = (
        "Name,HBL#,Email,Start time,Category,Description\n"
        "COSCO,HBL1,a@b.com,2026-08-01 09:00,Human Error,Needs check\n"
        ",MISSINGHBL,skip@x.com,2026-08-01 09:00,Human Error,\n"
    ).encode("utf-8")
    first = client.post(
        "/api/cases/import",
        data={"file": (BytesIO(csv_bytes), "cases.csv")},
        content_type="multipart/form-data",
    )
    assert first.status_code == 200
    body = first.get_json()
    assert body["imported"] == 1
    assert body["skipped"] == 1
    listed = client.get("/api/cases?q=HBL1").get_json()["data"]
    assert listed[0]["email"] == "a@b.com"
    case_id = listed[0]["id"]
    second = client.post(
        "/api/cases/import",
        data={"file": (BytesIO(csv_bytes), "cases.csv")},
        content_type="multipart/form-data",
    )
    assert second.get_json()["imported"] == 0
    assert second.get_json()["skipped"] >= 1
    updated = client.put(f"/api/cases/{case_id}", json={
        "status": "reviewed",
        "category": "Defects",
        "description": "Checked",
        "name": "HACKED",
        "email": "evil@x.com",
        "hbl": "NOPE",
    })
    assert updated.status_code == 200
    data = updated.get_json()["data"]
    assert data["name"] == "COSCO"
    assert data["email"] == "a@b.com"
    assert data["hbl"] == "HBL1"
    assert data["status"] == "reviewed"
    assert data["category"] == "Defects"
    assert data["description"] == "Checked"
    rejected = client.put(f"/api/cases/{case_id}", json={"category": "Not a real category"})
    assert rejected.status_code == 400


def test_case_import_xlsx():
    from io import BytesIO

    from openpyxl import Workbook

    migrate()
    from app import app

    client = app.test_client()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Name", "HBL#", "Start time", "Category", "Description"])
    sheet.append(["ONE", "HBLX", "2026-08-02 10:00", "System Enhancements", "Hub mismatch"])
    buffer = BytesIO()
    workbook.save(buffer)
    result = client.post(
        "/api/cases/import",
        data={"file": (BytesIO(buffer.getvalue()), "cases.xlsx")},
        content_type="multipart/form-data",
    )
    assert result.status_code == 200
    assert result.get_json()["imported"] == 1
    row = client.get("/api/cases?q=HBLX").get_json()["data"][0]
    assert row["name"] == "ONE"
    assert row["category"] == "System Enhancements"
    assert row["description"] == "Hub mismatch"


def test_case_template_download():
    migrate()
    from app import app

    client = app.test_client()
    response = client.get("/api/cases/template")
    assert response.status_code == 200
    body = response.get_data(as_text=True)
    assert "Category,Description" in body
    assert "Human Error" in body


def test_leave_name_mapping_and_half_day():
    from io import BytesIO

    from openpyxl import Workbook

    from services.leave import LEAVE_PEOPLE, replace_leave_people_from_workbook

    migrate()
    from app import app

    client = app.test_client()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Name Mapping"
    sheet.append(["Handled By", "Name"])
    sheet.append(["wenjie.yan@maersk.com", "Wenjie Yan"])
    sheet.append(["qing.huang@lns.maersk.com", "Qing Huang"])
    sheet.append(["jeff.yang@lns.maersk.com", "Jeff Yang"])
    sheet.append(["ailsa.he@lns.maersk.com", "Ailsa He"])
    sheet.append(["celia.liu@lns.maersk.com", "Celia Liu"])
    sheet.append(["Human Error", None])
    buffer = BytesIO()
    workbook.save(buffer)
    result, error = replace_leave_people_from_workbook(buffer.getvalue())
    assert error is None
    listed = client.get("/api/leave-people").get_json()["data"]
    names = [row["name"] for row in listed]
    assert set(LEAVE_PEOPLE).issubset(names)
    assert "Celia Liu" in names
    assert "Jane Li" in names
    assert "Jeff Yang" not in names
    assert "Ailsa He" not in names
    created = client.post("/api/leave-plans", json={
        "person": "Wenjie Yan",
        "leaveDate": "2026-08-12",
        "leaveType": "half day",
        "status": "planned",
    })
    assert created.status_code == 201
    assert created.get_json()["data"]["leaveType"] == "half_day"
    logs = client.get("/api/activity-logs?module=leave").get_json()["data"]
    assert any("Wenjie Yan" in (row.get("summary") or "") and "Half day" in (row.get("summary") or "") for row in logs)
    edited = client.put(f"/api/leave-plans/{created.get_json()['data']['id']}", json={
        "person": "Jane Li",
        "leaveDate": "2026-08-12",
        "leaveType": "annual",
        "status": "confirmed",
    })
    assert edited.status_code == 200
    removed = client.delete(f"/api/leave-plans/{created.get_json()['data']['id']}")
    assert removed.status_code == 200
    logs = client.get("/api/activity-logs?module=leave").get_json()["data"]
    assert len(logs) >= 3
    rejected = client.post("/api/leave-plans", json={
        "person": "Wenjie Yan",
        "leaveDate": "2026-08-13",
        "leaveType": "vacation",
        "status": "planned",
    })
    assert rejected.status_code == 400


def test_replace_cases_from_gca_workbook():
    from io import BytesIO

    from openpyxl import Workbook

    from db import get_connection
    from services.cases import replace_cases_from_gca_workbook

    migrate()
    from app import app

    client = app.test_client()
    client.post("/api/cases", json={"category": "Defects", "description": "seed to replace"})
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "AreaFeedbackList"
    sheet.append([
        "Start time", "Name", "HBL#", "Wrongly identify field", "Incorrect", "Corrected",
        "GSC PIC", "Week", "Date", "Category", "Description", "Action",
    ])
    sheet.append([
        "2026-03-01 09:15:00", "COSCO", "SZX1234567", "POR", "Guangzhou", "Shenzhen",
        0, "#N/A", "2026-03-01", "Human error", "Wrong POR", "Fixed",
    ])
    sheet.append([
        "2026-03-02 10:00:00", "Test user", "HBLTEST", "POD", "A", "B",
        "PIC", 12, "2026-03-02", "Test", "Ignore this", "",
    ])
    sheet.append([
        "2026-03-03 11:00:00", "ONE", "HBLFALLBACK", "Commodity", "Toys", "Electronics",
        "Ann", 13, "2026-03-03", "Defects", "", "",
    ])
    eur = workbook.create_sheet("AreaFeedbackList_EUR")
    eur.append(["HBL#", "Category", "Description", "Date"])
    eur.append(["EURHBL001", "System Enhancements", "EUR lane note", "2026-03-04"])
    buffer = BytesIO()
    workbook.save(buffer)
    result, error = replace_cases_from_gca_workbook(buffer.getvalue())
    assert error is None
    assert result["imported"] == 3
    assert result["skipped"] == 1
    listed = client.get("/api/cases").get_json()["data"]
    assert len(listed) == 3
    by_hbl = {row["hbl"]: row for row in listed}
    assert "seed to replace" not in {row["description"] for row in listed}
    assert by_hbl["SZX1234567"]["category"] == "Human Error"
    assert by_hbl["SZX1234567"]["gscPic"] == ""
    assert by_hbl["SZX1234567"]["week"] == ""
    assert by_hbl["HBLFALLBACK"]["description"] == "Commodity. Toys → Electronics"
    assert by_hbl["EURHBL001"]["category"] == "System Enhancements"


def test_lcl_import_requires_xlsx():
    from io import BytesIO
    from zipfile import ZipFile

    migrate()
    from app import app

    client = app.test_client()
    missing = client.post("/api/lcl/import")
    assert missing.status_code == 400
    csv_file = client.post(
        "/api/lcl/import",
        data={"file": (BytesIO(b"not excel"), "lcl.csv")},
        content_type="multipart/form-data",
    )
    assert csv_file.status_code == 400
    buffer = BytesIO()
    with ZipFile(buffer, "w") as archive:
        archive.writestr("xl/workbook.xml", "<workbook/>")
    fake = client.post(
        "/api/lcl/import",
        data={"file": (BytesIO(buffer.getvalue()), "lcl.xlsx")},
        content_type="multipart/form-data",
    )
    assert fake.status_code == 400
    assert "raw sheet" in fake.get_json()["message"].lower() or "pivot cache" in fake.get_json()["message"].lower()


def test_lcl_import_raw_sheet_headers():
    from io import BytesIO

    from openpyxl import Workbook

    migrate()
    from app import app

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Raw"
    sheet.append([
        "Shipment ID",
        "Job Branch",
        "Dest Ctry",
        "Weight",
        "Volume",
        "Dimension",
        "Chargeable",
        "Shipment Controlling Party Name",
        "Direction",
        "Month Name",
        "Count of Bosch",
        "Year",
        "Year Month",
        "Country Full Name",
    ])
    sheet.append([
        "EXP1", "SZ1", "DE", 100, 1.2, "1 x 100 x 80 x 60", 1.2,
        "ACME", "Export", "January", 1, 2026, "2026-01", "Germany",
    ])
    sheet.append([
        "IMP1", "SH1", "US", 80, 0.8, "1 x 80 x 60 x 50", 0.8,
        "BETA", "Import", "February", 0, 2026, "2026-02", "United States",
    ])
    payload = BytesIO()
    workbook.save(payload)
    client = app.test_client()
    imported = client.post(
        "/api/lcl/import",
        data={"file": (BytesIO(payload.getvalue()), "LCL_Volume_Analysis.xlsx")},
        content_type="multipart/form-data",
    )
    assert imported.status_code == 200
    body = imported.get_json()
    assert body["success"] is True
    assert body["data"]["exportCount"] == 1
    assert body["data"]["importCount"] == 1
    assert body["data"]["total"] == 2


def test_lcl_dimension_and_summary():
    from services.lcl import build_map, build_summary, clear_lcl_cache, parse_dimension
    from db import get_connection

    pieces, length, width, height = parse_dimension(" 2 x 120 x 80 x 60")
    assert pieces == 2
    assert length == 120
    migrate()
    with get_connection() as conn:
        conn.execute("DELETE FROM LclShipments")
        conn.execute("DELETE FROM LclImportMeta")
        conn.execute(
            """
            INSERT INTO LclShipments (
                ShipmentID, Direction, Year, MonthName, JobBranch, DestCtry, CountryName,
                Customer, IsBosch, Weight, Volume, DimensionRaw, Pieces, DimL, DimW, DimH, Chargeable
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("S1", "Export", "2026", "May", "SIN", "DE", "Germany", "BOSCH", 1, 100, 0.5, "1 x 0 x 0 x 0", 1, 0, 0, 0, 0.5),
        )
        conn.execute(
            """
            INSERT INTO LclShipments (
                ShipmentID, Direction, Year, MonthName, JobBranch, DestCtry, CountryName,
                Customer, IsBosch, Weight, Volume, DimensionRaw, Pieces, DimL, DimW, DimH, Chargeable
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("S2", "Export", "2026", "May", "SIN", "NL", "Netherlands", "PUMA", 0, 200, 1.5, "2 x 0 x 0 x 0", 2, 0, 0, 0, 1.5),
        )
    clear_lcl_cache()
    data = build_summary({"year": "2026"})
    assert data["kpis"]["shipments"] == 2
    assert data["kpis"]["avgVolume"] == 1.0
    assert data["byYearMonth"]["years"]
    assert data["byYearMonth"]["years"]
    mapped = build_map({"year": "2026"})
    points = mapped["points"] if isinstance(mapped, dict) else mapped
    codes = {item["iso2"] for item in points}
    assert "DE" in codes
    assert "NL" in codes
    assert mapped["arrows"]


def test_icb_csv_import_and_search():
    from io import BytesIO

    migrate()
    from app import app

    csv_bytes = (
        "Country,Location,CW1 Branch,CW1 UNLOCO Code (Consol),CW1 Group Code,CW1 Group Name,CW1 Agent Code,ICB Code\n"
        "Denmark ,Aarhus ,AAR,DKAAR,GR-AAR-FES,Denmark Export Sea (MAE),DAMDENAAR,DAMDENAAR\n"
        'México,Mexico,MX1,"MXLZC\nMXZLO",GR-MX1-FES,Mexico Station Export LCL (MAE),DAMLOGMEX,DAMLOGMEX\n'
        "Bangladesh,Dhaka,DAC,BDCGP,GR-DAC-FIS,Bangladesh Import Sea (MAE),APMGLODAC,APMGLODAC\n"
        "Chile,Santiago,SC1,CLSAI,GR-SC1-FES,Santiago Export Sea,MAELOGVAP,MAELOGVAP (as customer)\n"
    ).encode("utf-8")
    client = app.test_client()
    imported = client.post(
        "/api/icb/import",
        data={"file": (BytesIO(csv_bytes), "ICB.csv")},
        content_type="multipart/form-data",
    )
    assert imported.status_code == 200
    body = imported.get_json()
    assert body["success"] is True
    assert body["data"]["rowCount"] == 4
    aar = client.get("/api/icb?q=AAR").get_json()
    assert aar["pagination"]["total"] == 1
    assert aar["data"][0]["branch"] == "AAR"
    assert aar["data"][0]["icbCode"] == "DAMDENAAR"
    mexico = client.get("/api/icb?q=MXLZC").get_json()["data"][0]
    assert "MXLZC" in mexico["unloco"]
    assert "MXZLO" in mexico["unloco"]
    assert "\n" not in mexico["unloco"]
    inbound = client.get("/api/icb?q=Dhaka").get_json()["data"][0]
    assert inbound["direction"] == "import"
    chile = client.get("/api/icb?q=MAELOGVAP").get_json()["data"][0]
    assert chile["icbCode"] == "MAELOGVAP"
    assert "as customer" in chile["notes"]


def test_unloco_csv_import_and_search():
    from io import BytesIO

    migrate()
    from app import app

    csv_bytes = (
        "Rl PortName,UN Code,Rl Rn NKCountryCode,Country name,LCL Category,"
        "Is Ocean Primary Code,Rl IsActive,Rl IsSystem,Rl HasAirport,Rl HasSeaport,"
        "Rl HasRail,Rl HasRoad,Rl HasPost,Rl HasCustomsLodge,Rl HasUnload,Rl HasStore,"
        "Rl HasTerminal,Rl HasDischarge,Rl HasOutport,Rl HasBorderCrossing\n"
        "Lannion,FRLAI,FR,France,Door,TRUE,TRUE,TRUE,TRUE,TRUE,FALSE,FALSE,FALSE,FALSE,FALSE,FALSE,FALSE,FALSE,FALSE,FALSE\n"
        "Aarhus,DKAAR,DK,Denmark,Port,TRUE,TRUE,TRUE,FALSE,TRUE,FALSE,TRUE,FALSE,FALSE,FALSE,FALSE,FALSE,FALSE,FALSE,FALSE\n"
    ).encode("utf-8")
    client = app.test_client()
    imported = client.post(
        "/api/unlocode/import",
        data={"file": (BytesIO(csv_bytes), "UNLOCODE.csv")},
        content_type="multipart/form-data",
    )
    assert imported.status_code == 200
    body = imported.get_json()
    assert body["success"] is True
    assert body["data"]["rowCount"] == 2
    found = client.get("/api/unlocode?q=FRLAI").get_json()
    assert found["pagination"]["total"] == 1
    row = found["data"][0]
    assert row["unCode"] == "FRLAI"
    assert row["portName"] == "Lannion"
    airport = next(flag for flag in row["flags"] if flag["id"] == "hasAirport")
    seaport = next(flag for flag in row["flags"] if flag["id"] == "hasSeaport")
    road = next(flag for flag in row["flags"] if flag["id"] == "hasRoad")
    assert airport["on"] is True
    assert seaport["on"] is True
    assert road["on"] is False
    denmark = client.get("/api/unlocode?q=Aarhus").get_json()["data"][0]
    assert denmark["unCode"] == "DKAAR"
    by_country = client.get("/api/unlocode?q=France").get_json()
    assert by_country["pagination"]["total"] == 1
    assert by_country["data"][0]["unCode"] == "FRLAI"
    by_code = client.get("/api/unlocode?q=DK").get_json()
    assert by_code["data"][0]["unCode"] == "DKAAR"
    category_only = client.get("/api/unlocode?q=Door").get_json()
    assert category_only["pagination"]["total"] == 0


def test_gca_xlsx_import_summary_and_hbl_join():
    from datetime import datetime
    from io import BytesIO

    from openpyxl import Workbook

    migrate()
    from app import app

    workbook = Workbook()
    sz1 = workbook.active
    sz1.title = "SZ1"
    sz1.append(["Date", "#", "Order ID", "Booking ID", "Name", "Status", "Remark", "UID", "SCM", "HBL", "HBL#"])
    sz1.append([datetime(2026, 8, 1), 1, "OD1", "B1", "Alice", "Converted", None, "U1", "Bosch", None, "SZ10000001"])
    sz1.append([datetime(2026, 8, 2), 2, "OD2", "B2", "Bob", "Cancelled", "missing POR", "U2", None, None, "SZ10000002"])
    eur = workbook.create_sheet("SZ EUR")
    eur.append(["Date", "#", "Order ID", "Booking ID", "Name", "Status", "Remark", "UID", "SCM", "HBL", "HBL#"])
    eur.append([datetime(2026, 8, 1), 1, "OD3", "B3", "Cara", "Converted", None, "U3", None, None, "SZ10000003"])
    feedback = workbook.create_sheet("AreaFeedbackList")
    feedback.append([
        "Id", "Start time", "Email", "Name", "HBL#", "Wrongly identify field", "Incorrect ", "Corrected",
        "Cause of Error (Optional)", "GSC PIC", "Date", "Category",
    ])
    feedback.append([
        1, datetime(2026, 8, 1, 9, 0), "a@b.com", "Planner", "26SZ10000001",
        "[Party] - Shipper", "old", "new", "Misunderstanding", "Alice", datetime(2026, 8, 1), "Human error",
    ])
    feedback.append([
        2, datetime(2026, 8, 1, 9, 5), "a@b.com", "Planner", "TEST",
        "[Pre-fix Branch]", "SH1", "SZ1", "Test", 0, datetime(2026, 8, 1), "Test",
    ])
    eur_fb = workbook.create_sheet("AreaFeedbackList_EUR")
    eur_fb.append([
        "Id", "Name", "HBL#", "Wrongly identify field", "Incorrect ", "Corrected",
        "Cause of Error (Optional)", "GSC PIC", "Date", "Category",
    ])
    eur_fb.append([
        1, "Xie", "26SZ10000003", "[Service Information] - Incoterm", "FOB", "EXW",
        "Misunderstanding", "Cara", datetime(2026, 8, 1), "Defects",
    ])
    buffer = BytesIO()
    workbook.save(buffer)
    client = app.test_client()
    imported = client.post(
        "/api/gca/import",
        data={"file": (BytesIO(buffer.getvalue()), "gca.xlsx")},
        content_type="multipart/form-data",
    )
    assert imported.status_code == 200
    assert imported.get_json()["data"]["bookingCount"] == 3
    assert imported.get_json()["data"]["feedbackCount"] == 3
    summary = client.get("/api/gca/summary").get_json()["data"]
    assert summary["kpis"]["received"] == 3
    assert summary["kpis"]["converted"] == 2
    assert summary["kpis"]["cancelled"] == 1
    assert summary["kpis"]["feedbackCount"] == 2
    europe = client.get("/api/gca/summary?lane=europe").get_json()["data"]
    assert europe["kpis"]["received"] == 1
    bookings = client.get("/api/gca/bookings?q=OD1").get_json()["data"]
    assert bookings[0]["hbl"] == "SZ10000001"
    assert bookings[0]["feedbackCount"] == 1
    joined = client.get("/api/gca/feedback?q=SZ10000001").get_json()["data"]
    assert joined[0]["hblKey"] == "SZ10000001"
    assert joined[0]["orderId"] == "OD1"
    assert joined[0]["category"] == "Human Error"


if __name__ == "__main__":
    names = [name for name, value in list(globals().items()) if name.startswith("test_") and callable(value)]
    failed = 0
    for name in names:
        try:
            globals()[name]()
            print(f"PASS {name}")
        except Exception as error:
            failed += 1
            print(f"FAIL {name}: {error}")
    raise SystemExit(failed)
