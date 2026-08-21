import uuid
from datetime import datetime
from pathlib import Path

import mammoth
from flask import abort
from openpyxl import load_workbook

from config import ALLOWED_FILE_KINDS, MAX_UPLOAD_MB, PREVIEW_COLS, PREVIEW_ROWS, UPLOAD_DIR
from db import get_connection
from util import now_stamp


def ensure_upload_dir():
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    return UPLOAD_DIR


def file_kind_from_name(filename):
    suffix = Path(filename or "").suffix.lower()
    return ALLOWED_FILE_KINDS.get(suffix), suffix


def stored_path(stored_name):
    path = (ensure_upload_dir() / stored_name).resolve()
    if path.parent != ensure_upload_dir().resolve():
        abort(404)
    return path


def file_to_dict(row):
    return {
        "id": row["ID"],
        "originalName": row["OriginalName"],
        "kind": row["Kind"],
        "size": row["Size"],
        "uploadedAt": row["UploadedAt"],
        "updatedAt": row["UpdatedAt"],
    }


def cell_to_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def sanitize_preview_html(value):
    import re

    cleaned = re.sub(r"(?is)<script.*?>.*?</script>", "", value or "")
    cleaned = re.sub(r"(?is)<iframe.*?>.*?</iframe>", "", cleaned)
    return cleaned


def preview_docx(path):
    with path.open("rb") as handle:
        result = mammoth.convert_to_html(handle)
    return {
        "kind": "docx",
        "html": sanitize_preview_html(result.value),
        "messages": [str(item) for item in (result.messages or [])][:8],
    }


def preview_xlsx(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    try:
        for worksheet in workbook.worksheets:
            rows = []
            truncated = False
            for row_index, row in enumerate(worksheet.iter_rows(max_col=PREVIEW_COLS, values_only=True), start=1):
                if row_index > PREVIEW_ROWS:
                    truncated = True
                    break
                values = [cell_to_text(cell) for cell in row]
                if any(values):
                    rows.append(values)
            sheets.append({
                "name": worksheet.title,
                "rows": rows,
                "truncated": truncated,
            })
    finally:
        workbook.close()
    return {"kind": "xlsx", "sheets": sheets}


def save_bytes_as_file(original_name, data):
    from services.rag import index_file

    kind, suffix = file_kind_from_name(original_name)
    if not kind:
        return None, "Please upload a .docx or .xlsx file."
    if not data:
        return None, "The uploaded file is empty."
    if len(data) > MAX_UPLOAD_MB * 1024 * 1024:
        return None, f"File exceeds the {MAX_UPLOAD_MB} MB limit."
    display_name = Path(original_name).name.strip() or f"upload{suffix}"
    stored_name = f"{uuid.uuid4().hex}{suffix}"
    dest = stored_path(stored_name)
    tmp = dest.with_name(dest.name + ".tmp")
    tmp.write_bytes(data)
    stamp = now_stamp()
    try:
        with get_connection() as conn:
            cur = conn.execute(
                """
                INSERT INTO ToolkitFiles (OriginalName, StoredName, Kind, Size, UploadedAt, UpdatedAt)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (display_name, stored_name, kind, len(data), stamp, stamp),
            )
            file_id = cur.lastrowid
            tmp.replace(dest)
            row = conn.execute("SELECT * FROM ToolkitFiles WHERE ID=?", (file_id,)).fetchone()
            index_file(conn, row["ID"])
        return file_to_dict(row), None
    except Exception:
        if tmp.exists():
            tmp.unlink()
        if dest.exists() and dest.stat().st_size == len(data):
            pass
        raise
