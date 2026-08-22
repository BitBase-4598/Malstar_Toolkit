import json
import re
import sqlite3
import urllib.error
import urllib.request

import mammoth
from openpyxl import load_workbook

from config import (
    ASK_MAX_QUESTION,
    AZURE_OPENAI_API_KEY,
    AZURE_OPENAI_API_VERSION,
    AZURE_OPENAI_CHAT_DEPLOYMENT,
    AZURE_OPENAI_ENDPOINT,
    AZURE_OPENAI_TIMEOUT,
    PREVIEW_COLS,
    PREVIEW_ROWS,
    RAG_CHUNK_OVERLAP,
    RAG_CHUNK_SIZE,
    RAG_EXCERPT,
    RAG_TOP_K,
    llm_enabled,
)
from logging_util import audit
from services.files_store import cell_to_text, preview_docx, stored_path
from services.sops import load_sop
from util import now_stamp


def fts_available(conn):
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='RagChunksFts'"
    ).fetchone()
    return bool(row)


def touch_index_state(conn):
    count = conn.execute("SELECT COUNT(*) FROM RagChunks").fetchone()[0]
    conn.execute(
        """
        INSERT INTO RagIndexState (ID, LastIndexedAt, ChunkCount)
        VALUES (1, ?, ?)
        ON CONFLICT(ID) DO UPDATE SET
            LastIndexedAt=excluded.LastIndexedAt,
            ChunkCount=excluded.ChunkCount
        """,
        (now_stamp(), count),
    )


def chunk_text(text, size=RAG_CHUNK_SIZE, overlap=RAG_CHUNK_OVERLAP):
    cleaned = re.sub(r"\s+", " ", text or "").strip()
    if not cleaned:
        return []
    if len(cleaned) <= size:
        return [cleaned]
    pieces = []
    start = 0
    length = len(cleaned)
    while start < length:
        end = min(start + size, length)
        if end < length:
            space = cleaned.rfind(" ", start + size // 2, end)
            if space > start:
                end = space
        piece = cleaned[start:end].strip()
        if piece:
            pieces.append(piece)
        if end >= length:
            break
        start = max(end - overlap, start + 1)
    return pieces


def replace_source_chunks(conn, source_type, source_id, chunks, touch=True):
    conn.execute(
        "DELETE FROM RagChunks WHERE SourceType=? AND SourceID=?",
        (source_type, source_id),
    )
    stamp = now_stamp()
    for chunk in chunks:
        body = str(chunk.get("body") or "").strip()
        if not body:
            continue
        conn.execute(
            """
            INSERT INTO RagChunks (SourceType, SourceID, Title, Locator, Body, UpdatedAt)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                source_type,
                source_id,
                str(chunk.get("title") or "")[:300],
                str(chunk.get("locator") or "")[:120],
                body[:8000],
                stamp,
            ),
        )
    if touch:
        touch_index_state(conn)


def extract_docx_text(path):
    with path.open("rb") as handle:
        raw = mammoth.extract_raw_text(handle).value or ""
    text = re.sub(r"\s+", " ", raw).strip()
    if text:
        return text
    html = preview_docx(path).get("html") or ""
    return re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", html)).strip()


def extract_xlsx_chunks(path, title, file_id):
    workbook = load_workbook(path, read_only=True, data_only=True)
    chunks = []
    try:
        for worksheet in workbook.worksheets:
            for row_index, row in enumerate(
                worksheet.iter_rows(max_col=PREVIEW_COLS, values_only=True), start=1
            ):
                if row_index > PREVIEW_ROWS:
                    break
                values = [cell_to_text(cell).strip() for cell in row]
                values = [item for item in values if item]
                if not values:
                    continue
                body = f"{title} | {worksheet.title} row {row_index}: " + " | ".join(values)
                chunks.append({
                    "title": title,
                    "locator": f"{worksheet.title} row {row_index}",
                    "body": body,
                    "sourceId": file_id,
                })
    finally:
        workbook.close()
    return chunks


def index_sop(conn, sop_id, touch=True):
    data = load_sop(conn, sop_id)
    if not data:
        replace_source_chunks(conn, "sop", sop_id, [], touch=touch)
        return
    title = data["title"]
    chunks = []
    header = [f"SOP: {title}", f"Status: {data['status']}"]
    if data.get("owner"):
        header.append(f"Owner: {data['owner']}")
    if data.get("revision"):
        header.append(f"Revision: {data['revision']}")
    if data.get("purpose"):
        header.append(f"Purpose: {data['purpose']}")
    chunks.append({
        "title": title,
        "locator": "overview",
        "body": "\n".join(header),
    })
    for step in data.get("steps") or []:
        instruction = str(step.get("instruction") or "").strip()
        if not instruction:
            continue
        number = step.get("stepNumber")
        chunks.append({
            "title": title,
            "locator": f"step {number}",
            "body": f"SOP {title} step {number}: {instruction}",
        })
    replace_source_chunks(conn, "sop", sop_id, chunks, touch=touch)


def index_file(conn, file_id, touch=True):
    row = conn.execute("SELECT * FROM ToolkitFiles WHERE ID=?", (file_id,)).fetchone()
    if not row:
        replace_source_chunks(conn, "file", file_id, [], touch=touch)
        return
    title = row["OriginalName"]
    path = stored_path(row["StoredName"])
    if not path.is_file():
        replace_source_chunks(conn, "file", file_id, [], touch=touch)
        return
    chunks = []
    try:
        if row["Kind"] == "docx":
            text = extract_docx_text(path)
            for index, piece in enumerate(chunk_text(text), start=1):
                chunks.append({
                    "title": title,
                    "locator": f"section {index}",
                    "body": f"{title}: {piece}",
                })
        else:
            chunks = extract_xlsx_chunks(path, title, file_id)
    except Exception as error:
        audit(
            "file.index",
            "failure",
            resource_id=file_id,
            summary=f"{title}: {error}",
        )
        chunks = []
    replace_source_chunks(conn, "file", file_id, chunks, touch=touch)


def reindex_all(conn):
    conn.execute("DELETE FROM RagChunks")
    for row in conn.execute("SELECT ID FROM Sops").fetchall():
        index_sop(conn, row["ID"], touch=False)
    for row in conn.execute("SELECT ID FROM ToolkitFiles").fetchall():
        index_file(conn, row["ID"], touch=False)
    touch_index_state(conn)


def maybe_backfill_index(conn):
    count = conn.execute("SELECT COUNT(*) FROM RagChunks").fetchone()[0]
    if count:
        return
    files = conn.execute("SELECT COUNT(*) FROM ToolkitFiles").fetchone()[0]
    sops = conn.execute("SELECT COUNT(*) FROM Sops").fetchone()[0]
    if files or sops:
        reindex_all(conn)


def rag_status(conn):
    maybe_backfill_index(conn)
    state = conn.execute("SELECT LastIndexedAt, ChunkCount FROM RagIndexState WHERE ID=1").fetchone()
    sop_sources = conn.execute(
        "SELECT COUNT(DISTINCT SourceID) FROM RagChunks WHERE SourceType='sop'"
    ).fetchone()[0]
    file_sources = conn.execute(
        "SELECT COUNT(DISTINCT SourceID) FROM RagChunks WHERE SourceType='file'"
    ).fetchone()[0]
    chunk_count = state["ChunkCount"] if state else 0
    return {
        "chunkCount": chunk_count,
        "lastIndexedAt": state["LastIndexedAt"] if state else "",
        "llmEnabled": llm_enabled(),
        "sources": {
            "sops": sop_sources,
            "files": file_sources,
        },
    }


def build_fts_query(question):
    terms = re.findall(r"[A-Za-z0-9]{2,}", question or "")
    if terms:
        return " OR ".join(terms[:24])
    stripped = re.sub(r"[^\w\s]", " ", question or "", flags=re.UNICODE).strip()
    if stripped:
        return '"' + stripped.replace('"', "") + '"'
    return None


def search_chunks(conn, question, limit=RAG_TOP_K):
    query = build_fts_query(question)
    rows = []
    if query and fts_available(conn):
        try:
            rows = conn.execute(
                """
                SELECT c.ID, c.SourceType, c.SourceID, c.Title, c.Locator, c.Body
                FROM RagChunksFts
                JOIN RagChunks c ON c.ID = RagChunksFts.rowid
                WHERE RagChunksFts MATCH ?
                ORDER BY bm25(RagChunksFts)
                LIMIT ?
                """,
                (query, limit),
            ).fetchall()
        except sqlite3.OperationalError:
            rows = []
    if rows:
        return rows
    tokens = re.findall(r"[A-Za-z0-9]{2,}", question or "")
    if not tokens:
        leftover = (question or "").strip()
        tokens = [leftover] if leftover else []
    if not tokens:
        return []
    clauses = []
    params = []
    for token in tokens[:6]:
        like = f"%{token}%"
        clauses.append("(Title LIKE ? OR Locator LIKE ? OR Body LIKE ?)")
        params.extend([like, like, like])
    return conn.execute(
        f"""
        SELECT ID, SourceType, SourceID, Title, Locator, Body
        FROM RagChunks
        WHERE {' OR '.join(clauses)}
        ORDER BY ID DESC
        LIMIT ?
        """,
        params + [limit],
    ).fetchall()


def chunk_to_citation(row):
    body = str(row["Body"] or "").strip()
    if len(body) > RAG_EXCERPT:
        clipped = body[: RAG_EXCERPT - 1]
        space = clipped.rfind(" ")
        excerpt = (clipped[:space] if space > 80 else clipped) + "…"
    else:
        excerpt = body
    return {
        "sourceType": row["SourceType"],
        "sourceId": row["SourceID"],
        "title": row["Title"],
        "locator": row["Locator"],
        "excerpt": excerpt,
    }


def retrieve_only_answer(citations):
    if not citations:
        return (
            "No matching SOP or file excerpts were found. "
            "Try different words, or rebuild the index after uploading files or saving SOPs."
        )
    lines = [
        "Azure OpenAI is not configured, so here are the closest matches from Files and SOPs."
    ]
    for index, item in enumerate(citations, start=1):
        lines.append(f"[{index}] {item['title']} ({item['locator']})")
        lines.append(item["excerpt"])
    return "\n".join(lines)


def generate_answer(question, citations):
    if not citations:
        return (
            "I could not find this in the indexed Files and SOPs. "
            "Try different words, or rebuild the index after uploading files or saving SOPs."
        ), None
    if not llm_enabled():
        return retrieve_only_answer(citations), None
    numbered = []
    for index, item in enumerate(citations, start=1):
        numbered.append(
            f"[{index}] {item['sourceType'].upper()} \"{item['title']}\" ({item['locator']})\n{item['excerpt']}"
        )
    payload = {
        "messages": [
            {
                "role": "system",
                "content": (
                    "You answer questions using only the provided SOP and file excerpts. "
                    "If the excerpts do not contain the answer, say you could not find it "
                    "in the indexed Files and SOPs. Cite sources as [1], [2], and so on. "
                    "Do not invent procedures or file contents."
                ),
            },
            {
                "role": "user",
                "content": f"Question: {question}\n\nSources:\n" + "\n\n".join(numbered),
            },
        ],
        "temperature": 0.1,
        "max_tokens": 800,
    }
    url = (
        f"{AZURE_OPENAI_ENDPOINT}/openai/deployments/{AZURE_OPENAI_CHAT_DEPLOYMENT}"
        f"/chat/completions?api-version={AZURE_OPENAI_API_VERSION}"
    )
    request_obj = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "api-key": AZURE_OPENAI_API_KEY,
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request_obj, timeout=AZURE_OPENAI_TIMEOUT) as response:
            data = json.loads(response.read().decode("utf-8"))
        answer = (
            (data.get("choices") or [{}])[0]
            .get("message", {})
            .get("content", "")
            .strip()
        )
        if answer:
            return answer, None
        return retrieve_only_answer(citations), "Azure OpenAI returned an empty answer."
    except urllib.error.HTTPError as error:
        detail = error.read().decode("utf-8", errors="replace")[:300]
        return retrieve_only_answer(citations), f"Azure OpenAI HTTP {error.code}: {detail}"
    except Exception as error:
        return retrieve_only_answer(citations), str(error)
