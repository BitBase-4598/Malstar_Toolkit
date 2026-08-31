import csv
from io import BytesIO, StringIO

from openpyxl import load_workbook

from config import CASES_MAX_IMPORT_ROWS, MAX_JSON_UPLOAD_MB
from db import get_connection
from services.files_store import (
    build_file_preview,
    cell_to_text,
    stored_path,
    validate_upload_bytes,
    write_upload_bytes,
)
from util import now_stamp

STATUSES = ("pending_review", "reviewed", "closed")

CATEGORIES = (
    "Human Error",
    "Commercial knowledge & Operation Process Rules Updates",
    "System Enhancements",
    "Defects",
    "Invalid Feedback",
    "Process ambiguity reclarification",
)

CATEGORY_ALIASES = {
    "human error": "Human Error",
    "commercial knowledge & operation process rules updates": "Commercial knowledge & Operation Process Rules Updates",
    "system enhancements": "System Enhancements",
    "defects": "Defects",
    "invalid feedback": "Invalid Feedback",
    "process ambiguity reclarification": "Process ambiguity reclarification",
}

TEMPLATE_CSV = (
    "Category,Description\n"
    "Human Error,Example: POR was read as Guangzhou\n"
)

IMPORT_HEADERS = {
    "start time": "startTime",
    "completion time": "completionTime",
    "email": "email",
    "name": "name",
    "hbl": "hbl",
    "hbl no": "hbl",
    "hbl number": "hbl",
    "wrongly identified": "wronglyIdentified",
    "incorrect": "incorrect",
    "corrected": "corrected",
    "cause of error": "causeOfError",
    "received date": "receivedDate",
    "adjusted hbl": "adjustedHbl",
    "gsc pic": "gscPic",
    "week": "week",
    "date": "date",
    "category": "category",
    "description": "description",
    "action": "action",
}

TEXT_FIELDS = (
    ("startTime", "StartTime"),
    ("completionTime", "CompletionTime"),
    ("email", "Email"),
    ("name", "Name"),
    ("hbl", "HBL"),
    ("wronglyIdentified", "WronglyIdentified"),
    ("incorrect", "Incorrect"),
    ("corrected", "Corrected"),
    ("causeOfError", "CauseOfError"),
    ("receivedDate", "ReceivedDate"),
    ("adjustedHbl", "AdjustedHBL"),
    ("gscPic", "GscPic"),
    ("week", "Week"),
    ("date", "Date"),
    ("category", "Category"),
    ("description", "Description"),
    ("action", "Action"),
)


def case_file_to_dict(row):
    return {
        "id": row["ID"],
        "caseId": row["CaseID"],
        "originalName": row["OriginalName"],
        "kind": row["Kind"],
        "size": row["Size"],
        "uploadedAt": row["UploadedAt"],
    }


def parse_status(value, default="pending_review"):
    status = str(value or default).strip()
    if status not in STATUSES:
        return None, "Status must be Pending review, Reviewed, or Closed."
    return status, None


def parse_category(value, *, current=None, required=False):
    text = str(value or "").strip()
    if not text:
        if required:
            return None, "Category is required."
        return "", None
    mapped = CATEGORY_ALIASES.get(text.casefold(), text)
    if mapped in CATEGORIES or (current is not None and text == current):
        return mapped if mapped in CATEGORIES else text, None
    return None, "Category is not in the allowed list."


def empty_text_fields():
    return {key: "" for key, _column in TEXT_FIELDS}


def parse_case_create(data):
    data = data or {}
    category, error = parse_category(data.get("category"), required=True)
    if error:
        return None, error
    description = str(data.get("description") or "").strip()
    if not description:
        return None, "Description is required."
    status, error = parse_status(data.get("status"))
    if error:
        return None, error
    payload = empty_text_fields()
    payload["name"] = str(data.get("name") or "").strip()
    payload["hbl"] = str(data.get("hbl") or "").strip()
    payload["email"] = str(data.get("email") or "").strip()
    payload["category"] = category
    payload["description"] = description
    payload["status"] = status
    return payload, None


def parse_case_review(data, current):
    data = data or {}
    status, error = parse_status(data.get("status"), default=current.get("status") or "pending_review")
    if error:
        return None, error
    category, error = parse_category(data.get("category"), current=current.get("category") or "")
    if error:
        return None, error
    return {
        "status": status,
        "category": category,
        "description": str(data.get("description") or "").strip(),
    }, None


def _payload_column_values(payload):
    return [payload[key] for key, _column in TEXT_FIELDS]


def case_to_dict(row, file_count=0, files=None):
    data = {
        "id": row["ID"],
        "status": row["Status"],
        "startTime": row["StartTime"],
        "completionTime": row["CompletionTime"],
        "email": row["Email"],
        "name": row["Name"],
        "hbl": row["HBL"],
        "wronglyIdentified": row["WronglyIdentified"],
        "incorrect": row["Incorrect"],
        "corrected": row["Corrected"],
        "causeOfError": row["CauseOfError"],
        "receivedDate": row["ReceivedDate"],
        "adjustedHbl": row["AdjustedHBL"],
        "gscPic": row["GscPic"],
        "week": row["Week"],
        "date": row["Date"],
        "category": row["Category"],
        "description": row["Description"],
        "action": row["Action"],
        "createdAt": row["CreatedAt"],
        "updatedAt": row["UpdatedAt"],
        "fileCount": file_count,
    }
    if files is not None:
        data["files"] = files
    return data


def list_cases(conn, query=""):
    q = str(query or "").strip()
    join = """
        SELECT
            Cases.*,
            COALESCE(files.cnt, 0) AS FileCount
        FROM Cases
        LEFT JOIN (
            SELECT CaseID, COUNT(*) AS cnt FROM CaseFiles GROUP BY CaseID
        ) files ON files.CaseID = Cases.ID
    """
    if q:
        like = f"%{q}%"
        rows = conn.execute(
            f"""
            {join}
            WHERE CAST(Cases.ID AS TEXT) LIKE ?
               OR Cases.Name LIKE ?
               OR Cases.HBL LIKE ?
               OR Cases.Email LIKE ?
               OR Cases.Category LIKE ?
               OR Cases.Description LIKE ?
            ORDER BY Cases.UpdatedAt DESC, Cases.ID DESC
            """,
            (like, like, like, like, like, like),
        ).fetchall()
    else:
        rows = conn.execute(
            f"{join} ORDER BY Cases.UpdatedAt DESC, Cases.ID DESC"
        ).fetchall()
    return [case_to_dict(row, file_count=row["FileCount"]) for row in rows]


def load_case(conn, case_id):
    row = conn.execute("SELECT * FROM Cases WHERE ID=?", (case_id,)).fetchone()
    if not row:
        return None
    files = [
        case_file_to_dict(item)
        for item in conn.execute(
            "SELECT * FROM CaseFiles WHERE CaseID=? ORDER BY ID DESC",
            (case_id,),
        ).fetchall()
    ]
    return case_to_dict(row, file_count=len(files), files=files)


def create_case(conn, payload):
    stamp = now_stamp()
    created = str(payload.get("createdAt") or stamp)
    updated = str(payload.get("updatedAt") or created)
    columns = ", ".join(["Status"] + [column for _key, column in TEXT_FIELDS] + ["CreatedAt", "UpdatedAt"])
    placeholders = ", ".join(["?"] * (3 + len(TEXT_FIELDS)))
    cur = conn.execute(
        f"INSERT INTO Cases ({columns}) VALUES ({placeholders})",
        [payload["status"], *_payload_column_values(payload), created, updated],
    )
    return load_case(conn, cur.lastrowid)


def update_case_review(conn, case_id, payload):
    conn.execute(
        """
        UPDATE Cases
        SET Status=?, Category=?, Description=?, UpdatedAt=?
        WHERE ID=?
        """,
        (
            payload["status"],
            payload["category"],
            payload["description"],
            now_stamp(),
            case_id,
        ),
    )
    return load_case(conn, case_id)


def set_case_status(conn, case_id, status):
    conn.execute(
        "UPDATE Cases SET Status=?, UpdatedAt=? WHERE ID=?",
        (status, now_stamp(), case_id),
    )
    return load_case(conn, case_id)


def case_file_disk_names(conn, case_id):
    return [
        row["StoredName"]
        for row in conn.execute(
            "SELECT StoredName FROM CaseFiles WHERE CaseID=?",
            (case_id,),
        ).fetchall()
    ]


def unlink_stored_files(stored_names):
    for name in stored_names:
        path = stored_path(name)
        if path.is_file():
            path.unlink()


def save_case_file(case_id, original_name, data):
    meta, error = validate_upload_bytes(original_name, data)
    if error:
        return None, error
    stamp = now_stamp()
    try:
        with get_connection() as conn:
            existing = conn.execute("SELECT ID FROM Cases WHERE ID=?", (case_id,)).fetchone()
            if not existing:
                return None, "Case not found"
            write_upload_bytes(meta["stored_name"], data)
            cur = conn.execute(
                """
                INSERT INTO CaseFiles (CaseID, OriginalName, StoredName, Kind, Size, UploadedAt)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    case_id,
                    meta["display_name"],
                    meta["stored_name"],
                    meta["kind"],
                    meta["size"],
                    stamp,
                ),
            )
            row = conn.execute("SELECT * FROM CaseFiles WHERE ID=?", (cur.lastrowid,)).fetchone()
            conn.execute(
                "UPDATE Cases SET UpdatedAt=? WHERE ID=?",
                (stamp, case_id),
            )
        return case_file_to_dict(row), None
    except Exception:
        dest = stored_path(meta["stored_name"])
        if dest.is_file():
            dest.unlink()
        raise


def load_case_file(conn, case_id, file_id):
    row = conn.execute(
        "SELECT * FROM CaseFiles WHERE ID=? AND CaseID=?",
        (file_id, case_id),
    ).fetchone()
    if not row:
        return None, None
    path = stored_path(row["StoredName"])
    return row, path


def preview_case_file(row, path, case_id):
    return build_file_preview(
        row["Kind"],
        path,
        f"/api/cases/{case_id}/files/{row['ID']}/content",
        case_file_to_dict(row),
    )


def decode_json_upload(raw):
    import base64

    try:
        payload = base64.b64decode(str(raw or ""), validate=False)
    except Exception:
        return None, "The file data is not valid base64."
    if len(payload) > MAX_JSON_UPLOAD_MB * 1024 * 1024:
        return None, f"JSON uploads are limited to {MAX_JSON_UPLOAD_MB} MB. Use a smaller file."
    return payload, None


def case_heading(record):
    if not record:
        return ""
    category = record.get("category") or "Feedback"
    return f"#{record['id']} · {category}"


def normalize_import_header(value):
    text = str(value or "").strip().lower()
    text = text.replace("#", " ").replace("_", " ").replace("-", " ")
    return " ".join(text.split())


def map_import_headers(headers):
    mapping = {}
    for header in headers:
        key = IMPORT_HEADERS.get(normalize_import_header(header))
        if key:
            mapping[key] = header
    return mapping


def import_duplicate_key(payload):
    return (
        str(payload.get("category") or "").strip(),
        str(payload.get("description") or "").strip(),
        str(payload.get("startTime") or "").strip(),
    )


def row_values_to_payload(raw):
    payload = empty_text_fields()
    for key in payload:
        payload[key] = str(raw.get(key) or "").strip()
    payload["status"] = "pending_review"
    return payload


def parse_csv_import_rows(data):
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames:
        return None, "CSV header row is missing."
    mapping = map_import_headers(reader.fieldnames)
    if "category" not in mapping or "description" not in mapping:
        return None, "Missing Category or Description column."
    rows = []
    for index, row in enumerate(reader, start=2):
        if not any(str(value or "").strip() for value in row.values()):
            continue
        raw = {key: row.get(header, "") for key, header in mapping.items()}
        rows.append((index, row_values_to_payload(raw)))
    return rows, None


def parse_xlsx_import_rows(data):
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        iterator = sheet.iter_rows(values_only=True)
        header_row = next(iterator, None)
        if not header_row:
            return None, "Excel header row is missing."
        headers = [cell_to_text(cell) for cell in header_row]
        mapping = map_import_headers(headers)
        if "category" not in mapping or "description" not in mapping:
            return None, "Missing Category or Description column."
        index_by_key = {key: headers.index(header) for key, header in mapping.items()}
        rows = []
        for index, row in enumerate(iterator, start=2):
            values = [cell_to_text(cell) for cell in row]
            if not any(str(value or "").strip() for value in values):
                continue
            raw = {}
            for key, col_index in index_by_key.items():
                raw[key] = values[col_index] if col_index < len(values) else ""
            rows.append((index, row_values_to_payload(raw)))
        return rows, None
    finally:
        workbook.close()


def import_cases_file(filename, data):
    name = str(filename or "").strip() or "cases"
    suffix = name.rsplit(".", 1)[-1].lower() if "." in name else ""
    if suffix not in ("csv", "xlsx"):
        return None, "Please select a .xlsx or .csv file."
    if not data:
        return None, "The uploaded file is empty."
    if suffix == "csv":
        rows, error = parse_csv_import_rows(data)
    else:
        rows, error = parse_xlsx_import_rows(data)
    if error:
        return None, error
    if len(rows) > CASES_MAX_IMPORT_ROWS:
        return None, f"Import is limited to {CASES_MAX_IMPORT_ROWS} rows."
    if not rows:
        return None, "The file contains no data rows."

    imported = 0
    skipped = 0
    errors = []
    with get_connection() as conn:
        seen = {
            (
                str(row["Category"] or "").strip(),
                str(row["Description"] or "").strip(),
                str(row["StartTime"] or "").strip(),
            )
            for row in conn.execute("SELECT Category, Description, StartTime FROM Cases").fetchall()
        }
        for index, payload in rows:
            category, error = parse_category(payload.get("category"), required=True)
            description = str(payload.get("description") or "").strip()
            if error or not description:
                skipped += 1
                errors.append({
                    "row": index,
                    "message": error or "Category and Description are required.",
                })
                continue
            payload["category"] = category
            payload["description"] = description
            key = import_duplicate_key(payload)
            if key in seen:
                skipped += 1
                continue
            create_case(conn, payload)
            seen.add(key)
            imported += 1
    return {
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:100],
    }, None


def _fallback_description(item):
    from services.gca import cell_text

    wrongly = cell_text(item.get("wronglyIdentified"))
    incorrect = cell_text(item.get("incorrect"))
    corrected = cell_text(item.get("corrected"))
    bits = []
    if wrongly:
        bits.append(wrongly)
    if incorrect and corrected:
        bits.append(f"{incorrect} → {corrected}")
    elif incorrect:
        bits.append(incorrect)
    elif corrected:
        bits.append(corrected)
    return ". ".join(bits)


def gca_feedback_row_to_case_payload(item):
    from services.gca import SKIP_FEEDBACK_CATEGORIES, cell_date, cell_text, normalize_category

    category = normalize_category(item.get("category"))
    if category.casefold() in SKIP_FEEDBACK_CATEGORIES:
        return None
    mapped, error = parse_category(category, required=True)
    if error:
        return None
    description = cell_text(item.get("description")) or _fallback_description(item)
    if not description:
        return None
    payload = empty_text_fields()
    payload["startTime"] = cell_text(item.get("startTime"))
    payload["completionTime"] = cell_text(item.get("completionTime"))
    payload["email"] = cell_text(item.get("email"))
    payload["name"] = cell_text(item.get("name"))
    payload["hbl"] = cell_text(item.get("hbl"))
    payload["wronglyIdentified"] = cell_text(item.get("wronglyIdentified"))
    payload["incorrect"] = cell_text(item.get("incorrect"))
    payload["corrected"] = cell_text(item.get("corrected"))
    payload["causeOfError"] = cell_text(item.get("cause") or item.get("causeOfError"))
    payload["receivedDate"] = cell_date(item.get("receivedDate"))
    payload["adjustedHbl"] = cell_text(item.get("adjustedHbl"))
    payload["gscPic"] = cell_text(item.get("gscPic"))
    payload["week"] = cell_text(item.get("week"))
    payload["date"] = cell_date(item.get("date")) or cell_date(item.get("receivedDate"))
    payload["category"] = mapped
    payload["description"] = description
    payload["action"] = cell_text(item.get("action"))
    payload["status"] = "pending_review"
    created = payload["startTime"] or payload["date"] or payload["receivedDate"]
    if created:
        payload["createdAt"] = created
    return payload


def replace_cases_from_gca_workbook(data):
    from services.gca import FEEDBACK_HEADERS, FEEDBACK_SHEETS, read_sheet_rows

    if not data:
        return None, "No file was uploaded."
    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return None, "The file is not a valid Excel workbook."
    payloads = []
    skipped = 0
    found = []
    try:
        for sheet_name in FEEDBACK_SHEETS:
            rows, error = read_sheet_rows(workbook, sheet_name, FEEDBACK_HEADERS)
            if error:
                continue
            found.append(sheet_name)
            for item in rows:
                payload = gca_feedback_row_to_case_payload(item)
                if payload is None:
                    skipped += 1
                    continue
                payloads.append(payload)
    finally:
        workbook.close()
    if not found:
        return None, "Workbook is missing Area feedback sheets."
    if not payloads:
        return None, "Workbook contains no Area feedback rows to import."
    with get_connection() as conn:
        stored = [
            row["StoredName"]
            for row in conn.execute("SELECT StoredName FROM CaseFiles").fetchall()
        ]
        conn.execute("DELETE FROM CaseFiles")
        conn.execute("DELETE FROM Cases")
        conn.execute("DELETE FROM sqlite_sequence WHERE name IN ('Cases', 'CaseFiles')")
        for payload in payloads:
            create_case(conn, payload)
    unlink_stored_files(stored)
    return {"imported": len(payloads), "skipped": skipped, "sheets": found}, None
