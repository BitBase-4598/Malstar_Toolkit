import re
from datetime import datetime, date, time
from io import BytesIO

from openpyxl import load_workbook

from db import get_connection
from util import now_stamp

BOOKING_SHEETS = {
    "SZ1": {"lane": "non-europe", "branch": "SZ1"},
    "SZ EUR": {"lane": "europe", "branch": "SZ EUR"},
}
FEEDBACK_SHEETS = {
    "AreaFeedbackList": {"lane": "non-europe"},
    "AreaFeedbackList_EUR": {"lane": "europe"},
}
BOOKING_HEADERS = {
    "date": "date",
    "#": "sequence",
    "order id": "orderId",
    "booking id": "bookingId",
    "name": "name",
    "status": "status",
    "remark": "remark",
    "uid": "uid",
    "scm": "scm",
    "hbl": "hbl",
    "hbl no": "hblNumber",
    "hbl number": "hblNumber",
    "category": "category",
}
FEEDBACK_HEADERS = {
    "hbl": "hbl",
    "hbl no": "hbl",
    "hbl number": "hbl",
    "wrongly identify field": "wronglyIdentified",
    "wrongly identified": "wronglyIdentified",
    "wrongly identified field": "wronglyIdentified",
    "incorrect": "incorrect",
    "corrected": "corrected",
    "cause of error optional": "cause",
    "cause of error": "cause",
    "cause": "cause",
    "gsc pic": "gscPic",
    "category": "category",
    "description": "description",
    "action": "action",
    "date": "date",
    "received date": "receivedDate",
    "adjusted hbl": "adjustedHbl",
    "adjusted hbl no": "adjustedHbl",
    "week": "week",
    "email": "email",
    "name": "name",
    "start time": "startTime",
    "completion time": "completionTime",
}
NAME_MAPPING_SHEET = "Name Mapping"
NAME_MAPPING_HEADERS = {
    "handled by": "email",
    "email": "email",
    "name": "name",
}
SKIP_FEEDBACK_CATEGORIES = {"test", ""}
NA_VALUES = {"", "-", "—", "–", "n/a", "#n/a", "na", "none", "null"}
CATEGORY_ALIASES = {
    "human error": "Human Error",
    "commercial knowledge & operation process rules updates": "Commercial knowledge & Operation Process Rules Updates",
    "system enhancements": "System Enhancements",
    "defects": "Defects",
    "invalid feedback": "Invalid Feedback",
    "process ambiguity reclarification": "Process ambiguity reclarification",
    "test": "Test",
}
MAX_SHEET_COLS = 30


def normalize_header(value):
    raw = str(value or "").strip()
    if raw == "#":
        return "#"
    text = raw.lower().replace("#", " no ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.hour or value.minute or value.second:
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date) and not isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return ""
    if isinstance(value, bool):
        return "Y" if value else ""
    if isinstance(value, int):
        if value == 0:
            return ""
        return str(value)
    if isinstance(value, float):
        if value == 0:
            return ""
        if value.is_integer():
            return str(int(value))
        return str(value)
    text = str(value).replace("\u00a0", " ").strip()
    if text.casefold() in NA_VALUES:
        return ""
    return text


def cell_date(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date) and not isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = cell_text(value)
    if not text:
        return ""
    match = re.match(r"^(\d{4}-\d{2}-\d{2})", text)
    return match.group(1) if match else ""


def hbl_key(value):
    text = cell_text(value).replace(" ", "")
    if not text:
        return ""
    return text[-10:]


def normalize_category(value):
    text = cell_text(value)
    if not text:
        return ""
    return CATEGORY_ALIASES.get(text.casefold(), text)


def is_ai_converted(status, remark, category):
    blob = " ".join([status, remark, category]).casefold()
    return "ai converted" in blob or "converted by ai" in blob


def map_headers(cells, catalog):
    mapping = {}
    for index, cell in enumerate(cells):
        key = catalog.get(normalize_header(cell))
        if key and key not in mapping:
            mapping[key] = index
    return mapping


def read_sheet_rows(workbook, name, catalog):
    if name not in workbook.sheetnames:
        return [], f"Missing sheet {name}"
    worksheet = workbook[name]
    iterator = worksheet.iter_rows(max_col=MAX_SHEET_COLS, values_only=True)
    header = next(iterator, None)
    if not header:
        return [], f"{name} has no header row."
    mapping = map_headers(header, catalog)
    rows = []
    for raw in iterator:
        if not raw or not any(cell_text(cell) for cell in raw):
            continue
        item = {}
        for key, index in mapping.items():
            item[key] = raw[index] if index < len(raw) else None
        rows.append(item)
    return rows, None


def parse_bookings(raw_rows, lane, branch):
    records = []
    for item in raw_rows:
        booking_date = cell_date(item.get("date"))
        order_id = cell_text(item.get("orderId"))
        booking_id = cell_text(item.get("bookingId"))
        hbl = cell_text(item.get("hblNumber")) or cell_text(item.get("hbl"))
        status = cell_text(item.get("status"))
        name = cell_text(item.get("name"))
        if not any((booking_date, order_id, booking_id, hbl, status, name)):
            continue
        remark = cell_text(item.get("remark"))
        category = cell_text(item.get("category"))
        records.append({
            "bookingDate": booking_date,
            "sequenceNo": cell_text(item.get("sequence")),
            "orderId": order_id,
            "bookingId": booking_id,
            "name": name,
            "status": status,
            "remark": remark,
            "uid": cell_text(item.get("uid")),
            "scm": cell_text(item.get("scm")),
            "hbl": hbl,
            "hblKey": hbl_key(hbl),
            "lane": lane,
            "branch": branch,
            "category": category,
            "isAi": 1 if is_ai_converted(status, remark, category) else 0,
        })
    return records


def parse_feedback(raw_rows, lane):
    records = []
    for item in raw_rows:
        hbl = cell_text(item.get("hbl"))
        wrongly = cell_text(item.get("wronglyIdentified"))
        category = normalize_category(item.get("category"))
        if not any((hbl, wrongly, category, cell_text(item.get("incorrect")))):
            continue
        gsc_pic = cell_text(item.get("gscPic"))
        records.append({
            "hbl": hbl,
            "hblKey": hbl_key(hbl) or hbl_key(item.get("adjustedHbl")),
            "adjustedHbl": cell_text(item.get("adjustedHbl")),
            "wronglyIdentified": wrongly,
            "incorrect": cell_text(item.get("incorrect")),
            "corrected": cell_text(item.get("corrected")),
            "cause": cell_text(item.get("cause")),
            "gscPic": gsc_pic,
            "category": category,
            "description": cell_text(item.get("description")),
            "action": cell_text(item.get("action")),
            "feedbackDate": cell_date(item.get("date")) or cell_date(item.get("receivedDate")),
            "week": cell_text(item.get("week")),
            "email": cell_text(item.get("email")),
            "name": cell_text(item.get("name")),
            "lane": lane,
        })
    return records


def import_gca_workbook(filename, data):
    if not data:
        return None, "No file was uploaded."
    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return None, "The file is not a valid Excel workbook."
    bookings = []
    feedback = []
    found = []
    name_rows = []
    try:
        for sheet_name, meta in BOOKING_SHEETS.items():
            rows, error = read_sheet_rows(workbook, sheet_name, BOOKING_HEADERS)
            if error:
                continue
            found.append(sheet_name)
            bookings.extend(parse_bookings(rows, meta["lane"], meta["branch"]))
        for sheet_name, meta in FEEDBACK_SHEETS.items():
            rows, error = read_sheet_rows(workbook, sheet_name, FEEDBACK_HEADERS)
            if error:
                continue
            found.append(sheet_name)
            feedback.extend(parse_feedback(rows, meta["lane"]))
        rows, error = read_sheet_rows(workbook, NAME_MAPPING_SHEET, NAME_MAPPING_HEADERS)
        if not error:
            found.append(NAME_MAPPING_SHEET)
            name_rows = rows
    finally:
        workbook.close()
    if not found:
        return None, "Workbook is missing SZ1, SZ EUR, and Area feedback sheets."
    if not bookings:
        return None, "Workbook contains no booking rows on SZ1 / SZ EUR."
    stamp = now_stamp()
    with get_connection() as conn:
        conn.execute("BEGIN IMMEDIATE")
        conn.execute("DELETE FROM GcaBookings")
        conn.execute("DELETE FROM GcaFeedback")
        conn.executemany(
            """
            INSERT INTO GcaBookings (
                BookingDate, SequenceNo, OrderId, BookingId, Name, Status, Remark,
                Uid, Scm, Hbl, HblKey, Lane, Branch, Category, IsAi
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    item["bookingDate"],
                    item["sequenceNo"],
                    item["orderId"],
                    item["bookingId"],
                    item["name"],
                    item["status"],
                    item["remark"],
                    item["uid"],
                    item["scm"],
                    item["hbl"],
                    item["hblKey"],
                    item["lane"],
                    item["branch"],
                    item["category"],
                    item["isAi"],
                )
                for item in bookings
            ],
        )
        conn.executemany(
            """
            INSERT INTO GcaFeedback (
                Hbl, HblKey, AdjustedHbl, WronglyIdentified, Incorrect, Corrected,
                Cause, GscPic, Category, Description, Action, FeedbackDate, Week,
                Email, Name, Lane
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    item["hbl"],
                    item["hblKey"],
                    item["adjustedHbl"],
                    item["wronglyIdentified"],
                    item["incorrect"],
                    item["corrected"],
                    item["cause"],
                    item["gscPic"],
                    item["category"],
                    item["description"],
                    item["action"],
                    item["feedbackDate"],
                    item["week"],
                    item["email"],
                    item["name"],
                    item["lane"],
                )
                for item in feedback
            ],
        )
        conn.execute(
            """
            INSERT INTO GcaImportMeta (ID, Filename, ImportedAt, BookingCount, FeedbackCount)
            VALUES (1, ?, ?, ?, ?)
            ON CONFLICT(ID) DO UPDATE SET
                Filename=excluded.Filename,
                ImportedAt=excluded.ImportedAt,
                BookingCount=excluded.BookingCount,
                FeedbackCount=excluded.FeedbackCount
            """,
            ((filename or "gca.xlsx")[:200], stamp, len(bookings), len(feedback)),
        )
    from services.leave import parse_name_mapping_rows, replace_leave_people

    people = parse_name_mapping_rows(name_rows)
    if people:
        replace_leave_people(people)
    return {
        "filename": (filename or "gca.xlsx")[:200],
        "importedAt": stamp,
        "bookingCount": len(bookings),
        "feedbackCount": len(feedback),
        "peopleCount": len(people),
        "sheets": found,
    }, None


def parse_lane(value):
    text = str(value or "").strip().casefold()
    if text in ("", "all"):
        return ""
    if text in ("europe", "eur", "sz eur"):
        return "europe"
    if text in ("non-europe", "noneurope", "non europe", "sz1"):
        return "non-europe"
    return ""


def parse_date_arg(value):
    text = str(value or "").strip()
    if not text:
        return "", None
    try:
        datetime.strptime(text, "%Y-%m-%d")
    except ValueError:
        return None, "Dates must be YYYY-MM-DD."
    return text, None


def _booking_filters(date_from, date_to, lane, alias="b"):
    clauses = []
    params = []
    if date_from:
        clauses.append(f"{alias}.BookingDate >= ?")
        params.append(date_from)
    if date_to:
        clauses.append(f"{alias}.BookingDate <= ?")
        params.append(date_to)
    if lane:
        clauses.append(f"{alias}.Lane = ?")
        params.append(lane)
    sql = (" AND " + " AND ".join(clauses)) if clauses else ""
    return sql, params


def _feedback_date_expr(alias="f"):
    return (
        f"COALESCE(NULLIF({alias}.FeedbackDate, ''), "
        f"(SELECT MIN(b.BookingDate) FROM GcaBookings b "
        f"WHERE b.HblKey != '' AND b.HblKey = {alias}.HblKey), '')"
    )


def _feedback_filters(date_from, date_to, lane):
    clauses = []
    params = []
    date_expr = _feedback_date_expr()
    if date_from:
        clauses.append(f"{date_expr} >= ?")
        params.append(date_from)
    if date_to:
        clauses.append(f"{date_expr} <= ?")
        params.append(date_to)
    if lane:
        clauses.append("f.Lane = ?")
        params.append(lane)
    sql = (" AND " + " AND ".join(clauses)) if clauses else ""
    return sql, params


def _pct(part, whole):
    if not whole:
        return 0.0
    return round(100.0 * part / whole, 1)


def _meta(conn):
    row = conn.execute("SELECT * FROM GcaImportMeta WHERE ID=1").fetchone()
    bounds = conn.execute(
        """
        SELECT MIN(BookingDate) AS DateMin, MAX(BookingDate) AS DateMax, COUNT(*) AS RowCount
        FROM GcaBookings
        WHERE BookingDate != ''
        """
    ).fetchone()
    return {
        "filename": row["Filename"] if row else "",
        "importedAt": row["ImportedAt"] if row else "",
        "bookingCount": row["BookingCount"] if row else 0,
        "feedbackCount": row["FeedbackCount"] if row else 0,
        "dateMin": bounds["DateMin"] or "",
        "dateMax": bounds["DateMax"] or "",
        "rowCount": bounds["RowCount"] or 0,
    }


def booking_to_dict(row):
    return {
        "id": row["ID"],
        "date": row["BookingDate"],
        "sequenceNo": row["SequenceNo"],
        "orderId": row["OrderId"],
        "bookingId": row["BookingId"],
        "name": row["Name"],
        "status": row["Status"],
        "remark": row["Remark"],
        "uid": row["Uid"],
        "scm": row["Scm"],
        "hbl": row["Hbl"],
        "hblKey": row["HblKey"],
        "lane": row["Lane"],
        "branch": row["Branch"],
        "category": row["Category"],
        "isAi": bool(row["IsAi"]),
        "feedbackCount": row["FeedbackCount"] if "FeedbackCount" in row.keys() else 0,
    }


def feedback_to_dict(row):
    return {
        "id": row["ID"],
        "hbl": row["Hbl"],
        "hblKey": row["HblKey"],
        "adjustedHbl": row["AdjustedHbl"],
        "wronglyIdentified": row["WronglyIdentified"],
        "incorrect": row["Incorrect"],
        "corrected": row["Corrected"],
        "cause": row["Cause"],
        "gscPic": row["GscPic"],
        "category": row["Category"],
        "description": row["Description"],
        "action": row["Action"],
        "date": row["ResolvedDate"] if "ResolvedDate" in row.keys() else row["FeedbackDate"],
        "week": row["Week"],
        "email": row["Email"],
        "name": row["Name"],
        "lane": row["ResolvedLane"] if "ResolvedLane" in row.keys() else row["Lane"],
        "orderId": row["OrderId"] if "OrderId" in row.keys() else "",
        "bookingStatus": row["BookingStatus"] if "BookingStatus" in row.keys() else "",
    }


def build_summary(date_from="", date_to="", lane=""):
    lane = parse_lane(lane)
    extra, params = _booking_filters(date_from, date_to, lane)
    fb_extra, fb_params = _feedback_filters(date_from, date_to, lane)
    with get_connection() as conn:
        meta = _meta(conn)
        if not date_from:
            date_from = meta["dateMin"]
        if not date_to:
            date_to = meta["dateMax"]
        extra, params = _booking_filters(date_from, date_to, lane)
        fb_extra, fb_params = _feedback_filters(date_from, date_to, lane)
        counts = conn.execute(
            f"""
            SELECT
                COUNT(*) AS received,
                SUM(CASE WHEN lower(Status)='converted' THEN 1 ELSE 0 END) AS converted,
                SUM(CASE WHEN lower(Status)='cancelled' THEN 1 ELSE 0 END) AS cancelled,
                SUM(CASE WHEN lower(Status) IN ('pending', 'processing') THEN 1 ELSE 0 END) AS pending,
                SUM(CASE WHEN IsAi=1 AND lower(Status)='converted' THEN 1 ELSE 0 END) AS aiConverted,
                SUM(CASE WHEN lower(Status)='converted' AND IsAi=0 THEN 1 ELSE 0 END) AS gscConverted
            FROM GcaBookings b
            WHERE 1=1 {extra}
            """,
            params,
        ).fetchone()
        by_day = conn.execute(
            f"""
            SELECT BookingDate AS label,
                   COUNT(*) AS received,
                   SUM(CASE WHEN lower(Status)='converted' THEN 1 ELSE 0 END) AS converted,
                   SUM(CASE WHEN lower(Status)='cancelled' THEN 1 ELSE 0 END) AS cancelled
            FROM GcaBookings b
            WHERE BookingDate != '' {extra}
            GROUP BY BookingDate
            ORDER BY BookingDate
            """,
            params,
        ).fetchall()
        by_status = conn.execute(
            f"""
            SELECT CASE WHEN Status='' THEN '(blank)' ELSE Status END AS label, COUNT(*) AS count
            FROM GcaBookings b
            WHERE 1=1 {extra}
            GROUP BY label
            ORDER BY count DESC, label
            """,
            params,
        ).fetchall()
        by_category = conn.execute(
            f"""
            SELECT CASE WHEN f.Category='' THEN '(blank)' ELSE f.Category END AS label, COUNT(*) AS count
            FROM GcaFeedback f
            WHERE lower(COALESCE(f.Category, '')) NOT IN ('test', '')
            {fb_extra}
            GROUP BY label
            ORDER BY count DESC, label
            """,
            fb_params,
        ).fetchall()
        feedback_count = conn.execute(
            f"""
            SELECT COUNT(*) FROM GcaFeedback f
            WHERE lower(COALESCE(f.Category, '')) NOT IN ('test', '')
            {fb_extra}
            """,
            fb_params,
        ).fetchone()[0]
        filtered = counts["received"] or 0
        converted = counts["converted"] or 0
        cancelled = counts["cancelled"] or 0
        pending = counts["pending"] or 0
        ai_converted = counts["aiConverted"] or 0
        gsc_converted = counts["gscConverted"] or 0
        kpis = {
            "received": filtered,
            "converted": converted,
            "cancelled": cancelled,
            "pending": pending,
            "gscConverted": gsc_converted,
            "aiConverted": ai_converted,
            "conversionRate": _pct(converted, filtered),
            "gscConversionRate": _pct(gsc_converted, filtered),
            "aiConversionRate": _pct(ai_converted, filtered),
            "cancelledRate": _pct(cancelled, filtered),
            "feedbackCount": feedback_count,
            "errorsPerBooking": round(feedback_count / converted, 3) if converted else 0,
        }
        meta["dateFrom"] = date_from
        meta["dateTo"] = date_to
        meta["filteredCount"] = filtered
        meta["lane"] = lane or "all"
    return {
        "kpis": kpis,
        "meta": meta,
        "series": {
            "byDay": [
                {
                    "label": row["label"],
                    "count": row["received"] or 0,
                    "converted": row["converted"] or 0,
                    "cancelled": row["cancelled"] or 0,
                }
                for row in by_day
            ],
            "byStatus": [{"label": row["label"], "count": row["count"]} for row in by_status],
            "byCategory": [{"label": row["label"], "count": row["count"]} for row in by_category],
        },
    }


def list_bookings(date_from="", date_to="", lane="", q="", page=1, page_size=50):
    lane = parse_lane(lane)
    page = max(int(page or 1), 1)
    page_size = min(max(int(page_size or 50), 1), 200)
    extra, params = _booking_filters(date_from, date_to, lane)
    q = cell_text(q)
    if q:
        like = f"%{q}%"
        extra += """
            AND (
                b.OrderId LIKE ? COLLATE NOCASE
                OR b.BookingId LIKE ? COLLATE NOCASE
                OR b.Name LIKE ? COLLATE NOCASE
                OR b.Hbl LIKE ? COLLATE NOCASE
                OR b.HblKey LIKE ? COLLATE NOCASE
                OR b.Scm LIKE ? COLLATE NOCASE
                OR b.Status LIKE ? COLLATE NOCASE
            )
        """
        params.extend([like] * 7)
    with get_connection() as conn:
        total = conn.execute(f"SELECT COUNT(*) FROM GcaBookings b WHERE 1=1 {extra}", params).fetchone()[0]
        rows = conn.execute(
            f"""
            SELECT b.*, (
                SELECT COUNT(*) FROM GcaFeedback f
                WHERE f.HblKey != '' AND f.HblKey = b.HblKey
                  AND (f.Lane = b.Lane OR f.Lane = '' OR b.Lane = '')
            ) AS FeedbackCount
            FROM GcaBookings b
            WHERE 1=1 {extra}
            ORDER BY b.BookingDate DESC, b.ID DESC
            LIMIT ? OFFSET ?
            """,
            params + [page_size, (page - 1) * page_size],
        ).fetchall()
    return {
        "data": [booking_to_dict(row) for row in rows],
        "pagination": {
            "page": page,
            "pageSize": page_size,
            "total": total,
            "totalPages": max((total + page_size - 1) // page_size, 1),
        },
    }


def list_feedback(date_from="", date_to="", lane="", q="", page=1, page_size=50):
    lane = parse_lane(lane)
    page = max(int(page or 1), 1)
    page_size = min(max(int(page_size or 50), 1), 200)
    extra, params = _feedback_filters(date_from, date_to, lane)
    q = cell_text(q)
    if q:
        like = f"%{q}%"
        extra += """
            AND (
                f.Hbl LIKE ? COLLATE NOCASE
                OR f.HblKey LIKE ? COLLATE NOCASE
                OR f.WronglyIdentified LIKE ? COLLATE NOCASE
                OR f.Category LIKE ? COLLATE NOCASE
                OR f.GscPic LIKE ? COLLATE NOCASE
                OR f.Name LIKE ? COLLATE NOCASE
                OR COALESCE((
                    SELECT b.OrderId FROM GcaBookings b
                    WHERE b.HblKey != '' AND b.HblKey = f.HblKey AND b.Lane = f.Lane
                    LIMIT 1
                ), (
                    SELECT b.OrderId FROM GcaBookings b
                    WHERE b.HblKey != '' AND b.HblKey = f.HblKey
                    LIMIT 1
                ), '') LIKE ? COLLATE NOCASE
            )
        """
        params.extend([like] * 7)
    from_sql = "FROM GcaFeedback f WHERE 1=1"
    with get_connection() as conn:
        total = conn.execute(f"SELECT COUNT(*) {from_sql} {extra}", params).fetchone()[0]
        rows = conn.execute(
            f"""
            SELECT f.*,
                   {_feedback_date_expr()} AS ResolvedDate,
                   f.Lane AS ResolvedLane,
                   COALESCE((
                       SELECT b.OrderId FROM GcaBookings b
                       WHERE b.HblKey != '' AND b.HblKey = f.HblKey AND b.Lane = f.Lane
                       LIMIT 1
                   ), (
                       SELECT b.OrderId FROM GcaBookings b
                       WHERE b.HblKey != '' AND b.HblKey = f.HblKey
                       LIMIT 1
                   ), '') AS OrderId,
                   COALESCE((
                       SELECT b.Status FROM GcaBookings b
                       WHERE b.HblKey != '' AND b.HblKey = f.HblKey AND b.Lane = f.Lane
                       LIMIT 1
                   ), (
                       SELECT b.Status FROM GcaBookings b
                       WHERE b.HblKey != '' AND b.HblKey = f.HblKey
                       LIMIT 1
                   ), '') AS BookingStatus
            {from_sql} {extra}
            ORDER BY ResolvedDate DESC, f.ID DESC
            LIMIT ? OFFSET ?
            """,
            params + [page_size, (page - 1) * page_size],
        ).fetchall()
    return {
        "data": [feedback_to_dict(row) for row in rows],
        "pagination": {
            "page": page,
            "pageSize": page_size,
            "total": total,
            "totalPages": max((total + page_size - 1) // page_size, 1),
        },
    }
